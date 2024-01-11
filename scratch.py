def find_max(num):
    max = num[0]
    for i in num:
        if i > max:
            max = i
    return max


def odd_num(num):
    i = 1
    while i != 100:
        print(i)
        i +=2


def even_num(num):
    i = 0
    while i != 100:
        print(i)
        i +=2

import random

class Dice:
    def roll(self):
        first=random.randint(1, 6)
        second=random.randint(1, 6)
        return first, second

dice = Dice()
print(dice.roll())

# excel table ------------------------------------------------------------------------------------------------------------------


import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def file_process(filename):
    wb = xl.load_workbook(filename)
    sheet=wb['Sheet1']


    for row in range(2, sheet.max_row +1):
        cell = sheet.cell(row, 3)
        corr_cell = cell.value * 0.9
        corr_cell_price = sheet.cell(row, 4)
        corr_cell_price.value = corr_cell

    values = Reference(sheet, 
            min_row=2, 
            max_row=18, 
            min_col=4, 
            max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')


    wb.save(filename)

filename = 'transactions.xlsx'
file_process(filename)




x=int(input())

i = 2
for i in range(2,x):
    if (i < 4 or i ==5 or i==7):
        print(i)
    i += 1
    if i > 7:
        if (i % 3!= 0 and i % 2 != 0 and i % 5 != 0 and i % 7!= 0):
            print(i)
        i += 1


def check_prime(num):
    if num <= 1:
        return True
    elif num ==2 or num == 3 or num == 5 or num == 7:
        return True
    elif num % 2 == 0 or num % 3 == 0 or num % 5 == 0 or num % 7 == 0:
        return False
    
x = int(input("Enter the number:  ")) 


if check_prime(x):
    print("Prime")
else:
    print("Not Prime")


#4-------------------------------------------------------------------------------
p=print
numb = input().split()

abs_=[abs(int(num)) for num in numb]

s=sum
p(s(abs_))  

#------------------------------------------------------------------------------------------------

def check_prime(num):
    if num <= 1:
        return True
    elif num ==2 or num == 3 or num == 5 or num == 7:
        return True
    elif num % 2 == 0 or num % 3 == 0 or num % 5 == 0 or num % 7 == 0:
        return False
    
while True:
        
    x = input("Enter the number(exit to quit):  ")
    if x.lower() == 'exit':
        break
    
    try:
        x = int(x)
        if check_prime(x):
            print("Prime")
        else:
            print("Not Prime")

    except ValueError:
        print("Invalid key entry! please try only numbers or type exit to quit")


#------------------------------------------------------------------------------------------------------------------------------------
items_to_process = ['item1', 'item2']

# Loop through the items and save their outputs in separate files
for item in items_to_process:
    # Generate some output based on the item (replace this with your logic)
    output = f""" I look forward 
    to hearing from 
    you soon"""
    
    # Define the filename for this item
    filename = f"{item}.doc"
    
    # Open the file in write mode and write the output
    with open(filename, "w") as file:
        file.write(output)

    print(f"Output for {item} has been saved to {filename}")

#-------------------------------------------------------------------------------------------------------------------------------------


import pandas as pd

# Define the Excel file path
excel_file = "Firms.xlsx"

# Load the Excel file into a pandas DataFrame
df = pd.read_excel(excel_file)

# Specify the column name you want to import as a list
column_name = "CompanyName"

# Check if the specified column exists in the DataFrame
if column_name in df.columns:
    # Get the column data as a list
    column_data = df[column_name].tolist()

    # Print the list
    print(column_data)
else:
    print(f"Column '{column_name}' does not exist in the Excel file.")

#-------------------------------------------------------------------------------------------------------------------------------------

import pandas as pd

# Define the Excel file path
excel_file = "Firms.xlsx"

# Load the Excel file into a pandas DataFrame
df = pd.read_excel(excel_file)

# Specify the column name you want to import as a list
column_name = "CompanyName"

column_data = df[column_name].tolist()



# Loop through the items and save their outputs in separate files
for item in column_data:
    # Generate some output based on the item (replace this with your logic)
    output = f""" I look forward 
    to hearing from 
    {item} soon"""
    
    # Define the filename for this item
    filename = f"{item}.doc"
    
    # Open the file in write mode and write the output
    with open(filename, "w") as file:
        file.write(output)

    print(f"Output for {item} has been saved to {filename}")



#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# Email configuration
sender_email = 'bbharathreddy37@gmail.com'
sender_password = 'kdzi'
recipient_email = 'iambharathbr@gmail.com'
subject = 'Applicaiton for JOB'

# Create a multipart message
msg = MIMEMultipart()
msg['From'] = sender_email
msg['To'] = recipient_email
msg['Subject'] = subject

# Attach the PDF file without reading it
pdf_file_path = 'CVUP.pdf'  # Replace with the path to your PDF file

# Create a MIMEApplication object for the PDF file
pdf_attachment = MIMEApplication(open(pdf_file_path, 'rb').read(), _subtype='pdf')
pdf_attachment.add_header('content-disposition', 'attachment', filename='Resume.pdf')

# Attach the PDF file to the email
msg.attach(pdf_attachment)

# Establish a connection with Gmail's SMTP server
try:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(sender_email, sender_password)

    # Send the email with the PDF attachment
    server.sendmail(sender_email, recipient_email, msg.as_string())
    print('Email with PDF attachment sent successfully!')
except Exception as e:
    print('An error occurred:', str(e))
finally:
    server.quit()

#-------------------------------------------------------------------------------------------------------------------------------------
print("Age caluclator")
x=int(input("Enter your DOB Day: "))
y=int(input("Enter your DOB Month: "))
z=int(input("Enter your DOB Year: "))

yr=2023-z
if x > 27:
    dy=31-abs(27-x)
else:
    dy=27-x
if y > 9 :
    yr -= 1
    
    mn=12-(y-9)
    y -= 1
else:
    mn = 9-y
    

print(f'''You are: {yr}years, 
      {mn} months, 
      {dy} days old
      ''')


#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
from cryptography.fernet import Fernet

key=Fernet.generate_key()
ciper_suite=Fernet(key)

message = "CjjT3ygwgXNAMm2n8yw83LCMaLL3EpdiU_1VxPcK1Jg="

encrypt_mes= ciper_suite.encrypt(message.encode())

print(encrypt_mes, key)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

from PIL import Image

def reduce_image_size(input_image_path, output_image_path, max_size=1024, quality=85):
    """Reduce image size while maintaining aspect ratio.
    
    Args:
        input_image_path (str): Path to the input image file.
        output_image_path (str): Path to save the reduced size image.
        max_size (int): Maximum size (in kilobytes) for the output image.
        quality (int): Quality of the output image (0 to 100).
    """
    with Image.open(input_image_path) as img:
        # Calculate new dimensions while maintaining aspect ratio
        width_percent = (max_size / float(img.size[0]))
        new_width = int(img.size[0] * width_percent)
        new_height = int(img.size[1] * width_percent)
        
        # Resize the image
        resized_img = img.resize((new_width, new_height), Image.ANTIALIAS)
        
        # Save the resized image with specified quality
        resized_img.save(output_image_path, quality=quality)
        
# Example usage
input_image_path = 'input_image.jpg'  # Replace this with the path to your input image file
output_image_path = 'output_image.jpg'  # Replace this with the desired output image path

reduce_image_size(input_image_path, output_image_path, max_size=1024, quality=85)
#-----------------------------------------------------------------------------------------------------------------------------------------

from sklearn.tree import DecisionTreeRegressor
from sklearn.model_selection import train_test_split

# Generate some sample data
x = [i for i in range(0,10000000, 2)]
y = [j**3 for j in x]

# Split the data into training and testing sets
x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.2, random_state=42)

# Create a DecisionTreeRegressor model
model = DecisionTreeRegressor()

# Fit the model to the training data
model.fit([[x_val] for x_val in x_train], y_train)

# Get user input
user_input = int(input('Enter a number: '))

# Predict the cubic value for the user input
predicted_value = model.predict([[user_input]])
print(f'Predicted value: {predicted_value[0]}')


#---------------------------------------------------------------------------------------------------------------------------------

print("Enter thhe size of the arrray")
num=int(input())
a=[]
for i in range(num):
    a.append(input())
print("Enter the value to delete")
val=input()
if val in a:
    a.remove(val)
    print('The new array')
    for i in range(num-1):
        print(a[i], end=" ")
else:
    print("Entered value doesnt exit")

    #----------------------------------------------------------------------------------------------------------------------------------
import sys
arr=[10, 22, 38, 27, 11]
tmp=0
print(arr)
for i in range(0, len(arr)):
    for j in range(i,len(arr)):
        if arr[i]> arr[j]:
            tmp=arr[i]
            arr[i]=arr[j]
            arr[j]=tmp

print(arr)
print(sys.getsizeof(arr))

#--------------------------------------------------------------------------------------------------------------------------------------
class Node:
    def __init__(self,data):
        self.data=data
        self.next=None

class Ssl:
    def __init__(self):
        self.head=None

    def traversal(self):
        if self.head is None:
            print('SSL is empty')
        
        else:
            a=self.head
            while a is not None:
                print(a.data, end=" ")
            
                a=a.next

n1=Node(8)
ssl=Ssl()
ssl.head=n1
n2=Node(7)
n1.next=n2
n3=Node(5)
n2.next=n3

n4=Node(9)
n3.next=n4

n5=Node(3)
n4.next=n5
ssl.traversal()


    #-------------------------------------------------------------------------------------------------------------------------------------

def flood_fill(matrix, row, col):
    if row < 0 or col < 0 or row >= 3 or col >= 3 or matrix[row][col] != 1:
        return
    
    matrix[row][col] = 2
    
    # Recur for top, bottom, left, and right neighbors
    flood_fill(matrix, row-1, col)
    flood_fill(matrix, row+1, col)
    flood_fill(matrix, row, col-1)
    flood_fill(matrix, row, col+1)

# Example 3x3 matrix with elements 1 and 0
matrix = [[1, 1, 0],
          [1, 1, 0],
          [1, 1, 1]]

# Select a position with 1 to start the flood fill
row, col = 2, 2

# Perform flood fill
flood_fill(matrix, row, col)

# Print the modified matrix
for row in matrix:
    print(row)

#-----------------------------------------------------------------------------------------------
import numpy as np
from timeit import default_timer as timer
a=np.random.randn(1000)
b=np.random.randn(1000)
A=list(a)
B=list(b)
T=1000

def dot1():
    dot =0 
    for i in range(len(A)):
        dot += A[i]*B[i]
    return dot

def dot2():
    return np.dot(a,b)

start=timer()
for t in range(T):
    dot1()
end=timer()
t1=end-start

start=timer()
for t in range(T):
    dot2()
end=timer()
t2=end-start

print(t1)
print(t2)

#--------------------------------------------------------------------------------------------------------

import numpy as np  

a=[]
i=0
while i<100:
    print(i)
    i+=2
    a.append(i)
print(a)

b=np.array(a)
c=b.reshape((5,10))
print(c)

#------------------------------------------------------------------------------------------------------------------------


import numpy as np
import  time
import sys

s=100


L1=range(s)
L2=range(s)

A1=np.arange(s)
A2=np.arange(s)

start=time.time()
result=[(x+y) for x,y in zip(L1,L2)]
print(result)
print((time.time()-start)*1000)

start=time.time()
result=A1+A2
print(result)

print((time.time()-start)*1000)

#----------------------------------------------------------------------------------------------------------------------------------------

import numpy as np
import matplotlib.pyplot as plt

x=np.arange(0,3*np.pi, 0.1)
print(x)
y= np.sin(x)

plt.plot(x,y)
plt.show()

#-------------------------------------#----------------------------------------------------------------------------------------------------------------------------------------

def WordsToNumber(input_text):
    dic={'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5,
         'six': 6, 'seven': 7, 'eight': 8, 'nine': 9, 'zero': 0}
    
    words=input_text.split()
    result=[]
    i=0
    while i <len(words):
        word=words[i]
        if word=='triple':
            num=dic.get(words[i+1], None)
            if num is not None:
                result.extend([num]*3)
            i +=2
        elif word=='double':
            num=dic.get(words[i+1], None)
            if num is not None:
                result.extend([num]*2)
            i +=2
        else:
            result.append(dic.get(word, None))
            i+=1

    return result

input_text = "five one zero six triple eight nine six four"

output=WordsToNumber(input_text)

res=int(''.join(map(str,output)))
print(res)

#----------------------------------------------------------------------------------------------------------------------------------------

import pandas as pd
import numpy as np
dates=pd.date_range('20230909', periods=6)
df=pd.DataFrame(np.random.rand(6,4), index=dates, columns=list("ABCD"))
print(df)

df=pd.DataFrame({'A':1.0,
                 'B':pd.Timestamp('20230909'),
                 'C':pd.Series(1, index=list(range(4)), dtype='float32'),
                 'D':np.array([3]*4, dtype='int32'),
                 'E':pd.Categorical(['Test','train','test', 'train']),
                 'F':'hello'})

print(df)

#----------------------------------------------------------------------------------------------------------------------------------------

import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score

test_data=pd.read_csv('test.csv')
train_data=pd.read_csv('train.csv')

test_data['Age'].fillna(test_data['Age'].median(), inplace=True)
train_data['Age'].fillna(train_data['Age'].median(), inplace=True)


test_data['Sex']=test_data['Sex'].map({'female':0, 'male':1})
train_data['Sex']=train_data['Sex'].map({'female':0, 'male':1})

features=['Pclass','Sex','Age','SibSp','Parch']
X=train_data[features]
y=train_data['Survived']

X_train, X_val, y_train, y_val=train_test_split(X,y, test_size=0.2, random_state=60)
model=RandomForestClassifier(n_estimators=100, random_state=60)
model.fit(X_train, y_train)

predictions=model.predict(X_val)

acuracy=accuracy_score(y_val,predictions)
print('Accuarcy: ', acuracy)

test_features=test_data[features]
test_predict=model.predict(test_features)


test_data['Survived']=test_predict

submission=test_data[['PassengerId','Survived']]
submission.to_csv('submission.csv',index=False)

#--------------------------------------------------------------------------------------------------------------------------------------------------

from datasets import load_dataset
import pandas as pd

# Load the dataset
dataset = load_dataset("squad_v2")

# Convert the dataset to pandas DataFrames
dataframes = {}
for split in dataset.keys():
    dataframes[split] = pd.DataFrame(dataset[split])

# Save the pandas DataFrames as a dictionary
dataframes_dict = {split: dataframe.to_dict(orient='records') for split, dataframe in dataframes.items()}

# Save the dictionary as a JSON file
import json
with open("dataset.json", "w") as json_file:
    json.dump(dataframes_dict, json_file)

#--------------------------------------------------------------------------------------------------------------------------------------------------

import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, classification_report
import joblib

# Load your dataset
data = pd.read_excel('validsquad.xlsx')  # Replace 'your_dataset.xlsx' with your actual file path

# Features: 'context' and 'question'
features = ['context', 'question']
X = data[features]

# Target variable: 'answers'
y = data['answers']

# Convert text data to numerical features using TF-IDF vectorization
tfidf_vectorizer = TfidfVectorizer()
X_tfidf = tfidf_vectorizer.fit_transform(X['context'] + ' ' + X['question'])

# Split the data into training and validation sets
X_train, X_val, y_train, y_val = train_test_split(X_tfidf, y, test_size=0.2, random_state=42)

# Create and train the RandomForestClassifier
model = RandomForestClassifier(n_estimators=30, random_state=42)

# Monitoring Training Progress
model.fit(X_train, y_train)

# Evaluate on Validation Data
predictions_val = model.predict(X_val)
accuracy = accuracy_score(y_val, predictions_val)
classification_rep = classification_report(y_val, predictions_val)

# Save the trained model to a file
joblib.dump(model, 'squad_modelvalid1.pkl')

# Print Training Results
print(f'Accuracy on validation data: {accuracy:.2f}')
print('Classification Report on validation data:')
print(classification_rep)

from sklearn.feature_extraction.text import TfidfVectorizer
import joblib

# Load the pre-trained RandomForestClassifier model
loaded_model = joblib.load('squad_modelvalid1.pkl')

# Input context and question manually
context = input("Enter context: ")
question = input("Enter question: ")

# Convert text data to numerical features using TF-IDF vectorization
tfidf_vectorizer = TfidfVectorizer()
input_tfidf = tfidf_vectorizer.transform([context + ' ' + question])

# Make predictions using the loaded model
prediction = loaded_model.predict(input_tfidf)

# Print the prediction
print("Predicted answer:", prediction[0])

#--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as pd
import tensorflow as tf
from tensorflow import keras
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
import json


# Replace 'your_excel_file.xlsx' with the path to your Excel file and 'Sheet1' with the sheet name
excel_file_path = 'validsqua.xlsx'
sheet_name = 'Sheet1'

# Read the Excel file and extract a specific column into a list
try:
    # Read the Excel file
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    
    # Replace 'Column_Name' with the name of the column you want to extract
    column_name1 ='context'
    column_name2= 'question'
    column_name3= 'answers'
    
    # Extract the specified column into a list
    column_data1 = df[column_name1].tolist()
    column_data2 = df[column_name2].tolist()
    column_data3 = df[column_name3].tolist()

    
    print("Data from  column in Excel file")

except Exception as e:
    print(f"An error occurred: {e}")



# Sample dataset (replace this with your own dataset)
contexts = column_data1  # List of context texts
questions = column_data2  # List of questions
answers = column_data3  # List of corresponding answers

# Tokenize texts
tokenizer = Tokenizer()
tokenizer.fit_on_texts(contexts + questions + answers)


# Convert tokenizer configuration to a dictionary
tokenizer_config = tokenizer.get_config()

# Save tokenizer to a JSON file
with open('tokenizer.json', 'w', encoding='utf-8') as json_file:
    json.dump(tokenizer_config, json_file)


# Convert texts to sequences
context_sequences = tokenizer.texts_to_sequences(contexts)
question_sequences = tokenizer.texts_to_sequences(questions)
answer_sequences = tokenizer.texts_to_sequences(answers)

# Pad sequences for uniform length
max_sequence_length = 100  # Set your desired sequence length
context_sequences = pad_sequences(context_sequences, maxlen=max_sequence_length, padding='post')
question_sequences = pad_sequences(question_sequences, maxlen=max_sequence_length, padding='post')
answer_sequences = pad_sequences(answer_sequences, maxlen=max_sequence_length, padding='post')

# Define the model architecture
input_context = tf.keras.layers.Input(shape=(max_sequence_length,))
input_question = tf.keras.layers.Input(shape=(max_sequence_length,))

# Create embedding layers for context and question
embedding_dim = 100  # Set your desired embedding dimension
context_embedding = tf.keras.layers.Embedding(input_dim=len(tokenizer.word_index) + 1, output_dim=embedding_dim)(input_context)
question_embedding = tf.keras.layers.Embedding(input_dim=len(tokenizer.word_index) + 1, output_dim=embedding_dim)(input_question)

# Define your model architecture (you can customize this as needed)
# For example, you can use LSTM layers, attention mechanisms, etc.

# Concatenate context and question embeddings
merged = tf.keras.layers.Concatenate(axis=-1)([context_embedding, question_embedding])

# Add your layers for processing the concatenated embeddings
# For example, you can add LSTM layers, attention mechanisms, etc.

# Output layer
output = tf.keras.layers.Dense(len(tokenizer.word_index) + 1, activation='softmax')(merged)

# Create the model
model = tf.keras.models.Model(inputs=[input_context, input_question], outputs=output)

# Compile the model
model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy'])

# Train the model
model.fit([context_sequences, question_sequences], answer_sequences, epochs=10, batch_size=32)


model.save('bert.keras')

# Now, you can use this trained model to predict answers for new questions within a given context.
# To make predictions:
# 1. Tokenize and pad the new question and context.
# 2. Use model.predict() to get the answer sequence.
# 3. Decode the answer sequence back to text using the tokenizer.


#--------------------------------------------------------------------------------------------------------------------------------------------------

import pandas as pd
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
from keras.models import load_model
from keras.preprocessing.text import tokenizer_from_json
import json



# # Load the tokenizer and model
tokenizer = Tokenizer()
# tokenizer.fit_on_texts(['list of texts used for training'])  # Pass the list of texts used for training here


# Load tokenizer configuration from a JSON file
with open('tokenizer.json', 'r', encoding='utf-8') as json_file:
    tokenizer_config = json.load(json_file)

max_sequence_length = 100  # Set your desired sequence length
embedding_dim = 100  # Set your desired embedding dimension
# Create tokenizer from the loaded configuration
tokenizer = tokenizer_from_json(tokenizer_config)

# # Load the pre-trained tokenizer from a JSON file
# with open('tokenizer.json', 'r') as json_file:
#     tokenizer_config = json.load(json_file)
#     tokenizer = tokenizer_from_json(tokenizer_config)

# Load the pre-trained model
model = load_model('bert.keras')

# Example input text (multiple sentences or words)
input_context_text = input("Your input context text here. It can contain multiple sentences or words.")
input_question_text = input("Your input question text here.")

# Tokenize and pad the input text
input_context_sequences = tokenizer.texts_to_sequences([input_context_text])
input_question_sequences = tokenizer.texts_to_sequences([input_question_text])

input_context_padded = pad_sequences(input_context_sequences, maxlen=max_sequence_length, padding='post')
input_question_padded = pad_sequences(input_question_sequences, maxlen=max_sequence_length, padding='post')

# Perform prediction
predicted_answer_sequences = model.predict([input_context_padded, input_question_padded])

# Convert the predicted sequences back to text
predicted_answers = []
for answer_sequence in predicted_answer_sequences:
    predicted_answer = " ".join([list(tokenizer.word_index.keys())[i - 1] for i in answer_sequence if i > 0])
    predicted_answers.append(predicted_answer)

print("Predicted answers:", predicted_answers)


#--------------------------------------------------------------------------------------------------------------------------------------------------

import torch
from transformers import RobertaTokenizer, RobertaForQuestionAnswering
from torch.utils.data import DataLoader, TensorDataset
import pandas as pd

# Load pre-trained RoBERTa tokenizer and model
tokenizer = RobertaTokenizer.from_pretrained('roberta-base')
model = RobertaForQuestionAnswering.from_pretrained('roberta-base')

# Sample dataset (replace this with your own dataset)
contexts = ["Context 1 text.", "Context 2 text.", ...]  # List of context texts
questions = ["Question 1 text?", "Question 2 text?", ...]  # List of questions
answers = ["Answer 1 text.", "Answer 2 text.", ...]  # List of corresponding answers

# Tokenize and encode context-question pairs
encoded_inputs = tokenizer(contexts, questions, padding=True, truncation=True, return_tensors='pt')

# Convert answers to start and end token positions in the tokenized context
start_positions = []
end_positions = []

for answer, context in zip(answers, contexts):
    start_idx = context.find(answer)
    end_idx = start_idx + len(answer) - 1  # -1 to get the inclusive end position
    start_positions.append(tokenizer.encode(context[:start_idx], add_special_tokens=False).shape[1])
    end_positions.append(tokenizer.encode(context[:end_idx], add_special_tokens=False).shape[1] - 1)

start_positions = torch.tensor(start_positions)
end_positions = torch.tensor(end_positions)

# Create a PyTorch dataset
dataset = TensorDataset(encoded_inputs['input_ids'], encoded_inputs['attention_mask'], start_positions, end_positions)

# Data loader
batch_size = 8
dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=True)

# Training loop
num_epochs = 3  # Set your desired number of epochs
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
model.to(device)
optimizer = torch.optim.AdamW(model.parameters(), lr=5e-5)

for epoch in range(num_epochs):
    for batch in dataloader:
        input_ids, attention_mask, start_positions, end_positions = batch
        input_ids, attention_mask, start_positions, end_positions = input_ids.to(device), attention_mask.to(device), start_positions.to(device), end_positions.to(device)
        
        optimizer.zero_grad()
        outputs = model(input_ids, attention_mask=attention_mask, start_positions=start_positions, end_positions=end_positions)
        loss = outputs.loss
        loss.backward()
        optimizer.step()

# Now, you can use the trained model to get answers for questions within a given context
def get_answer(context, question):
    inputs = tokenizer(question, context, return_tensors='pt')  # Note the reversal of question and context
    inputs.to(device)
    outputs = model(**inputs)
    start_idx = torch.argmax(outputs.start_logits)
    end_idx = torch.argmax(outputs.end_logits)
    answer_tokens = inputs['input_ids'][0][start_idx:end_idx+1]
    answer = tokenizer.decode(answer_tokens)
    return answer

# Example usage
context = "Context 1 text."
question = "Question 1 text?"
predicted_answer = get_answer(context, question)
print("Predicted Answer:", predicted_answer)


#--------------------------------------------------------------------------------------------------------------------------------------------------

from transformers import GPT2Tokenizer, GPT2LMHeadModel

# Replace 'gpt-j' with the actual model name if it's different
model_name = 'EleutherAI/gpt-j-6B'  

# Load tokenizer and model
tokenizer = GPT2Tokenizer.from_pretrained(model_name)
model = GPT2LMHeadModel.from_pretrained(model_name)

# Example usage
text = ["earth", "dirt", "alligator"]
input_ids = tokenizer.encode(text, return_tensors='pt')

# Generate output
output = model.generate(input_ids, max_length=100, num_return_sequences=1)
generated_text = tokenizer.decode(output[0], skip_special_tokens=True)

print("Generated Text: ", generated_text)

#--------------------------------------------------------------------------------------------------------------------------------------------------

from transformers import BertTokenizer, BertForQuestionAnswering, AdamW
import torch
from torch.utils.data import DataLoader, TensorDataset

# Load pre-trained BERT tokenizer and model
tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
model = BertForQuestionAnswering.from_pretrained('bert-base-uncased')

# Define your dataset for question answering (input_ids, attention_mask, start_positions, end_positions)
# ...
# Assuming you have your data in the form of tensors
input_ids = ...  # Tensor containing input IDs
attention_masks = ...  # Tensor containing attention masks
start_positions = ...  # Tensor containing start positions for answers
end_positions = ...  # Tensor containing end positions for answers

# Create a TensorDataset from your data
dataset = TensorDataset(input_ids, attention_masks, start_positions, end_positions)

# Define batch size
batch_size = 32

# Create a DataLoader
dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=True)
# Set up optimizer
optimizer = AdamW(model.parameters(), lr=5e-5)

# Fine-tuning loop
num_epochs = 3
for epoch in range(num_epochs):
    for batch in dataloader:  # Iterate through your dataset batches
        optimizer.zero_grad()
        input_ids, attention_mask, start_positions, end_positions = batch
        outputs = model(input_ids, attention_mask=attention_mask, start_positions=start_positions, end_positions=end_positions)
        loss = outputs.loss
        loss.backward()
        optimizer.step()

# Save the fine-tuned model
model.save_pretrained('fine_tuned_bert_model')
tokenizer.save_pretrained('fine_tuned_bert_model')


# Define the context and question
context = "The Apollo program, also known as Project Apollo, was the third United States human spaceflight program carried out by the National Aeronautics and Space Administration (NASA)."
question = "What was the Apollo program?"

# Tokenize the input
inputs = tokenizer.encode_plus(question, context, return_tensors='pt', max_length=512, truncation=True)

# Get the answer
start_positions, end_positions = model(**inputs).values()

# Get the answer text
answer_start = torch.argmax(start_positions)
answer_end = torch.argmax(end_positions) + 1
answer = tokenizer.convert_tokens_to_string(tokenizer.convert_ids_to_tokens(inputs['input_ids'][0][answer_start:answer_end]))

print("Answer:", answer)


#--------------------------------------------------------------------------------------------------------------------------------------------------

from transformers import AutoModelForQuestionAnswering, AutoTokenizer, pipeline
import pandas as pd
from transformers import AutoModelForQuestionAnswering, AutoTokenizer
from torch.utils.data import DataLoader, Dataset
import torch
import ast


def predict ():
    model_name = "fine_tuned_roberta_model"

    # a) Get predictions
    nlp = pipeline('question-answering', model=model_name, tokenizer=model_name)
    QA_input = {
        'question': 'What accolade did Beyoncé receive from Forbes magazine in 2015?',
        'context': '''A self-described "modern-day feminist", Beyoncé creates songs that are often characterized by themes of love, relationships, and monogamy, as well as female sexuality and empowerment. On stage, her dynamic, highly choreographed performances have led to critics hailing her as one of the best entertainers in contemporary popular music. Throughout a career spanning 19 years, she has sold over 118 million records as a solo artist, and a further 60 million with Destiny's Child, making her one of the best-selling music artists of all time. She has won 20 Grammy Awards and is the most nominated woman in the award's history. The Recording Industry Association of America recognized her as the Top Certified Artist in America during the 2000s decade. In 2009, Billboard named her the Top Radio Songs Artist of the Decade, the Top Female Artist of the 2000s and their Artist of the Millennium in 2011. Time listed her among the 100 most influential people in the world in 2013 and 2014. Forbes magazine also listed her as the most powerful female musician of 2015.'''
    }
    res = nlp(QA_input)

    # b) Load model & tokenizer
    model = AutoModelForQuestionAnswering.from_pretrained(model_name)
    tokenizer = AutoTokenizer.from_pretrained(model_name)

    # {'text': ["In 2015, Forbes magazine recognized Beyoncé as the most powerful female musician, underscoring her impact and dominance in the music business."], 'answer_start': [556]}


    print(res)


def train():
    # Load your dataset from Excel
    df = pd.read_csv("F:\\python\\valid.csv",  encoding="latin1")

    # Parse the 'answers' column as a dictionary
    df['answers'] = df['answers'].apply(ast.literal_eval)

    class CustomDataset(Dataset):
        def __init__(self, data, tokenizer, max_sequence_length):
            self.data = data
            self.tokenizer = tokenizer
            self.max_sequence_length = max_sequence_length

        def __len__(self):
            return len(self.data)

        def __getitem__(self, idx):
            item = self.data.iloc[idx]
            context = item['context']
            question = item['question']
            
            # Check if the 'answers' field has the expected structure
            if 'answers' in item and 'answer_start' in item['answers'] and item['answers']['answer_start']:
                answer_start = item['answers']['answer_start'][0]  # Accessing the first element of the list
                answer_end = answer_start + len(item['answers']['text'][0])

                inputs = self.tokenizer(
                    question,
                    context,
                    truncation="only_second",
                    max_length=self.max_sequence_length,  # Specify the maximum sequence length
                    padding="max_length",  # Pad to the specified maximum length
                    return_tensors="pt"
                )

                input_ids = inputs.input_ids.squeeze()  # Remove the extra dimension
                attention_mask = inputs.attention_mask.squeeze()  # Remove the extra dimension
                start_positions = torch.tensor(answer_start)
                end_positions = torch.tensor(answer_end - 1)  # -1 because the end position is inclusive

                return {
                    "input_ids": input_ids,
                    "attention_mask": attention_mask,
                    "start_positions": start_positions,
                    "end_positions": end_positions
                }
            else:
                # Handle unexpected 'answers' field structure (skip this instance or handle it accordingly)
                return None  # Skip this instance for now

    # Load pre-trained RoBERTa model and tokenizer
    model_name = "deepset/roberta-base-squad2"
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = AutoModelForQuestionAnswering.from_pretrained(model_name)

    # Define the maximum sequence length for padding
    max_sequence_length = 128  # You can adjust this value based on your specific use case

    # Prepare DataLoader
    batch_size = 4
    train_dataset = CustomDataset(df, tokenizer, max_sequence_length)

    # Filter out None values from the dataset
    train_dataset = [data for data in train_dataset if data is not None]

    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)

    # Define training parameters
    optimizer = torch.optim.AdamW(model.parameters(), lr=1e-5)
    num_epochs = 3

    # Training loop
    for epoch in range(num_epochs):
        total_loss = 0
        total_correct_start = 0
        total_correct_end = 0
        total_instances = 0

        for batch in train_loader:
            optimizer.zero_grad()
            inputs = {
                "input_ids": batch["input_ids"],
                "attention_mask": batch["attention_mask"],
                "start_positions": batch["start_positions"],
                "end_positions": batch["end_positions"]
            }
            outputs = model(**inputs)
            loss = outputs.loss
            start_logits = outputs.start_logits
            end_logits = outputs.end_logits
            
            # Calculate accuracy for start positions
            pred_start_positions = torch.argmax(start_logits, dim=1)
            correct_start_positions = torch.sum(pred_start_positions == batch["start_positions"])
            
            # Calculate accuracy for end positions
            pred_end_positions = torch.argmax(end_logits, dim=1)
            correct_end_positions = torch.sum(pred_end_positions == batch["end_positions"])
            
            # Update total correct positions and total instances
            total_correct_start += correct_start_positions.item()
            total_correct_end += correct_end_positions.item()
            total_instances += len(batch["input_ids"])

            # Backpropagation and optimization
            loss.backward()
            optimizer.step()
            total_loss += loss.item()

        average_loss = total_loss / len(train_loader)
        start_accuracy = total_correct_start / total_instances
        end_accuracy = total_correct_end / total_instances

        print(f"Epoch {epoch+1}/{num_epochs}, Loss: {average_loss:.4f}, Start Accuracy: {start_accuracy:.4f}, End Accuracy: {end_accuracy:.4f}")

    # Save the fine-tuned model and tokenizer
    model.save_pretrained("fine_tuned_roberta_model")
    tokenizer.save_pretrained("fine_tuned_roberta_model")

    print("Fine-tuning completed. Model and tokenizer saved.")




train()
predict()

#--------------------------------------------------------------------------------------------------------------------------------------------------

import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

df=pd.read_csv('F:\\python\\HR_comma_sep.csv')
features = ['satisfaction_level', 'last_evaluation', 'average_montly_hours', 'Work_accident', 'promotion_last_5years']
target = 'left'

# Split the data into features and target variable
X = df[features]
y = df[target]
X_train,X_test,y_train, y_test = train_test_split(X,y,test_size=0.01, random_state=42)
model=LogisticRegression()

model.fit(X_train,y_train)
prediction = model.predict(X_test)
fig = plt.figure(figsize=(8, 6))
ax = fig.add_subplot(111, projection='3d')

ax.scatter(df['satisfaction_level'], df['last_evaluation'], df['average_montly_hours'], c=df['Work_accident'], marker='o')

ax.set_xlabel('Satisfaction Level')
ax.set_ylabel('Last Evaluation')
ax.set_zlabel('Average Monthly Hours')

plt.show()
print(prediction)
acc=model.score(X_test,y_test)
print(acc)


#--------------------------------------------------------------------------------------------------------------------------------------------------

import pandas as pd
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression

df=pd.read_csv("F:\python\CSV\homeprices.csv")

plt.xlabel('area')
plt.ylabel('price')
plt.scatter(df.area,df.price,color='red', marker='+')

model=LinearRegression()

model.fit(df[['area']],df.price)
s=model.predict([[5000]])
print(s)
d=pd.read_csv("F:\\python\\CSV\\areas.csv")
pre = model.predict(d)
d['Predictions'] = list(pre)
d.to_csv('predict.csv',index=False)
print(d)




#--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as pd
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
import warnings as wr

df=pd.read_csv("F:\python\CSV\hiring.csv")

# plt.xlabel('area')
# plt.ylabel('price')
# plt.scatter(df.area,df.price,color='red', marker='+')

model=LinearRegression()
df.bedrooms=df.bedrooms.fillna(df.bedrooms.median())
model.fit(df[['area','bedrooms','age']],df.price)
s=model.predict([[5000,3,25]])

print(s)



#--------------------------------------------------------------------------------------------------------------------------------------------------


import numpy as np

def gradient_desent(x,y):
    m_cur = b_cur = 0
    iterations = 1000
    n=len(x)
    learning_rate = 0.001

    for i in range(iterations):
        y_pred=m_cur*x + b_cur
        cost = (1/n)*sum([val**2 for val in (y-y_pred)])
        md= - (2/n)*sum(x*(y-y_pred))
        bd = - (2/n)*sum(y-y_pred)
        m_cur = m_cur - md*learning_rate
        b_cur = b_cur - bd*learning_rate

        print("m{} b{} cost{} iteration{}".format(m_cur,b_cur,cost,i)) 



x=np.array([1,2,3,4,5])
y=np.array([5,7,9,11,13])

gradient_desent(x,y)


#Gradient descent/sklearn LR-------------------------------------------------------------------------------------------------------------------------------------------------


import numpy as np
import pandas as pd
from sklearn.linear_model import LinearRegression
import math

def by_sklearn(x,y):
    df=pd.read_csv("test_scores.csv")
    d=LinearRegression()
    d.fit(df[['math']], df.cs)
    return d.coef_, d.intercept_



def by_gradient_desent(x,y):
    m_cur = b_cur = 0
    iterations = 100000
    cost_init = 0
    learning_rate = 0.0002
    n= len(x)
    

    for i in range(iterations):
        y_predict = m_cur*x +b_cur
        cost = (1/n)*sum([val**2 for val in (y-y_predict)])
        md = -(2/n)*sum(x*(y-y_predict))
        bd = -(2/n)*sum(y-y_predict)

        m_cur= m_cur - learning_rate*md
        b_cur= b_cur - learning_rate*bd

        if math.isclose(cost, cost_init, rel_tol = 1e-20):
            break
        cost_init = cost
        print('m{} b{} cost{} iter{}'.format(m_cur,b_cur,cost,iterations))
    


if __name__ == '__main__':
    df = pd.read_csv("test_scores.csv")
    x = np.array(df.math)
    y = np.array(df.cs)

    m, b = by_gradient_desent(x,y)
    print("Using gradient descent function: Coef {} Intercept {}".format(m, b))

    m_sklearn, b_sklearn = by_sklearn()
    print("Using sklearn: Coef {} Intercept {}".format(m_sklearn,b_sklearn))



#Dummies with LinearRegression--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as p
from sklearn.linear_model import LinearRegression
df=p.read_csv("F:\\python\\CSV\\exer\\homeprices.csv")
dummies = p.get_dummies(df['town'])
merge = p.concat([df,dummies],axis='columns')
fin = merge.drop(['town','robinsville'],axis='columns')
X= fin.drop(['price'], axis='columns')
y= fin['price']

m=LinearRegression()
m.fit(X,y)
prd = m.predict([[3700,1,0]])
print(m.coef_)
print(m.score(X,y))
print(prd)



#--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as p
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import OneHotEncoder
from sklearn.preprocessing import LabelEncoder
from sklearn.compose import ColumnTransformer


df=p.read_csv("F:\\python\\CSV\\exer\\homeprices.csv")
dummies = p.get_dummies(df['town'])
merge = p.concat([df,dummies],axis='columns')
fin = merge.drop(['town','robinsville'],axis='columns')
X= fin.drop(['price'], axis='columns')
y= fin['price']

m=LinearRegression()
m.fit(X,y)
prd = m.predict([[3700,0,1]])
print(m.coef_)
print(m.score(X,y))
print(prd)

le=LabelEncoder()
dfle= df
dfle['town']= le.fit_transform(df['town'])
X = dfle[['town', 'area']].values
y = dfle['price'].values
ct = ColumnTransformer([('town', OneHotEncoder(), [0])], remainder='passthrough')
X = ct.fit_transform(X)
X = X[:,1:]
m.fit(X,y)
predic = m.predict([[0, 1, 3700]])
print(predic)




#--------------------------------------------------------------------------------------------------------------------------------------------------


from sklearn.datasets import load_digits
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt
import seaborn as sn
from sklearn.linear_model import LogisticRegression
import joblib
digits = load_digits()
X_train, X_test, y_train, y_test = train_test_split(digits['data'], digits['target'], test_size = 0.1, random_state=42)
model = LogisticRegression(max_iter=100000)
model.fit(X_train, y_train)
pred = model.predict(X_test)
joblib.dump(model, 'LogicRegDigits.pk1')
# loaded_model = joblib.load('model_filename.pkl')

score=model.score(X_test,y_test)
print(pred)
print(score)


#--------------------------------------------------------------------------------------------------------------------------------------------------


from sklearn.datasets import load_digits
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt
import seaborn as sn
from sklearn.linear_model import LogisticRegression
import joblib

# Load the Digits dataset
digits = load_digits()
X_train, X_test, y_train, y_test = train_test_split(digits['data'], digits['target'], test_size=0.1, random_state=42)

# Create and train the Logistic Regression model
model = LogisticRegression(max_iter=10000)
model.fit(X_train, y_train)

# Save the trained model to a file using joblib
joblib.dump(model, 'LogisticRegDigits.pkl')

# Calculate and print the accuracy score
score = model.score(X_test, y_test)
print("Accuracy Score:", score)

# Make predictions on the test set
y_pred = model.predict(X_test)

# Generate the confusion matrix
cm = confusion_matrix(y_test, y_pred)

# Visualize the confusion matrix using a heatmap
plt.figure(figsize=(8, 6))
sn.heatmap(cm, annot=True, fmt='g', cmap='Blues')
plt.xlabel('Predicted')
plt.ylabel('Actual')
plt.title('Confusion Matrix')
plt.show()



#--------------------------------------------------------------------------------------------------------------------------------------------------



import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier

df = pd.read_csv("F:\\python\\CSV\\titanic.csv")
X = df[['SibSp','Parch','Age','Sex','Pclass']].copy()

y = df['Survived']


X['Sex'] = X['Sex'].map({'male': 0, 'female': 1})
X['Age'] = X['Age'].fillna(X['Age'].mean())

# X_train, X_test, y_train, y_test = train_test_split(X,y,test_size=0.2, random_state=42)
# modelLR = LogisticRegression()
# modelLR.fit(X_train,y_train)


# Assuming X contains your features and y contains corresponding labels

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.6, random_state=42)

# Logistic Regression
logreg = LogisticRegression()
logreg_scores = cross_val_score(logreg, X_train, y_train, cv=5, scoring='accuracy')
print("Logistic Regression Cross-Validation Accuracy: ", logreg_scores.mean())

# Decision Tree Classifier
dt_classifier = DecisionTreeClassifier()
dt_scores = cross_val_score(dt_classifier, X_train, y_train, cv=5, scoring='accuracy')
print("Decision Tree Classifier Cross-Validation Accuracy: ", dt_scores.mean())

# Train and evaluate the models on the test set
logreg.fit(X_train, y_train)
logreg_predictions = logreg.predict(X_test)
print("Logistic Regression Test Accuracy: ", accuracy_score(y_test, logreg_predictions))
print("Logistic Regression Test Precision: ", precision_score(y_test, logreg_predictions))
print("Logistic Regression Test Recall: ", recall_score(y_test, logreg_predictions))
print("Logistic Regression Test F1-Score: ", f1_score(y_test, logreg_predictions))

dt_classifier.fit(X_train, y_train)
dt_predictions = dt_classifier.predict(X_test)
print("Decision Tree Classifier Test Accuracy: ", accuracy_score(y_test, dt_predictions))
print("Decision Tree Classifier Test Precision: ", precision_score(y_test, dt_predictions))
print("Decision Tree Classifier Test Recall: ", recall_score(y_test, dt_predictions))
print("Decision Tree Classifier Test F1-Score: ", f1_score(y_test, dt_predictions))




#--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as pd
from sklearn.preprocessing import MinMaxScaler
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt

df = pd.read_csv("F:\\python\\CSV\\income.csv")

plt.scatter(df['Age'], df['Income($)'])
plt.show()

km = KMeans(n_clusters=3)
pred = km.fit_predict(df[['Age','Income($)']])
df['cluster'] = pred

df1 = df[df['cluster']==0]
df2 = df[df['cluster']==1]
df3 = df[df['cluster']==2]

plt.scatter(df1['Age'], df1['Income($)'], color = 'red')
plt.scatter(df2['Age'], df2['Income($)'], color = 'blue')
plt.scatter(df3['Age'], df3['Income($)'], color = 'green')
plt.show()
scale = MinMaxScaler()

scale.fit(df[['Age']])
df['Age'] = scale.transform(df[['Age']])

scale.fit(df[['Income($)']])
df['Income($)'] = scale.transform(df[['Income($)']])

km = KMeans(n_clusters=3)
pred = km.fit_predict(df[['Age','Income($)']])
df['cluster'] = pred
df1 = df[df['cluster']==0]
df2 = df[df['cluster']==1]
df3 = df[df['cluster']==2]

plt.scatter(df1['Age'], df1['Income($)'], color = 'red')
plt.scatter(df2['Age'], df2['Income($)'], color = 'blue')
plt.scatter(df3['Age'], df3['Income($)'], color = 'green')
plt.show()


sse = []
k_rng = range(1,10)
for k in k_rng:
    km = KMeans(n_clusters=k)
    km.fit(df[['Age','Income($)']])
    sse.append(km.inertia_)
    
plt.plot(k_rng,sse)
plt.show()



#--------------------------------------------------------------------------------------------------------------------------------------------------



import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import RandomForestClassifier

df = pd.read_csv("F:\\python\\CSV\\titanic.csv")
X = df[['SibSp','Parch','Age','Sex','Pclass']].copy()

y = df['Survived']


X['Sex'] = X['Sex'].map({'male': 0, 'female': 1})
X['Age'] = X['Age'].fillna(X['Age'].mean())

# X_train, X_test, y_train, y_test = train_test_split(X,y,test_size=0.2, random_state=42)
# modelLR = LogisticRegression()
# modelLR.fit(X_train,y_train)


# Assuming X contains your features and y contains corresponding labels

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

# Logistic Regression
logreg = LogisticRegression()
logreg_scores = cross_val_score(logreg, X_train, y_train, cv=5, scoring='accuracy')
print("Logistic Regression Cross-Validation Accuracy: ", logreg_scores.mean())

# Decision Tree Classifier
dt_classifier = DecisionTreeClassifier()
dt_scores = cross_val_score(dt_classifier, X_train, y_train, cv=5, scoring='accuracy')
print("Decision Tree Classifier Cross-Validation Accuracy: ", dt_scores.mean())


# Randm Forest Classifier
rf_classifier = RandomForestClassifier()
rf_scores = cross_val_score(rf_classifier, X_train, y_train, cv=5, scoring='accuracy')
print("Random forest Classifier Cross-Validation Accuracy: ", rf_scores.mean())


# Naive Bayes
nb_classifier = GaussianNB()
nb_scores = cross_val_score(nb_classifier, X_train, y_train, cv=5, scoring='accuracy')
print("Naive Bayes Cross-Validation Accuracy: ", nb_scores.mean())




# Train and evaluate the models on the test set
rf_classifier.fit(X_train, y_train)
rf_predictions = rf_classifier.predict(X_test)
print("Random forest Classifier Test Accuracy: ", accuracy_score(y_test, rf_predictions))
print("Random forest Classifier Test Precision: ", precision_score(y_test, rf_predictions))
print("Random forest Classifier Test Recall: ", recall_score(y_test, rf_predictions))
print("Random forest Classifier Test F1-Score: ", f1_score(y_test, rf_predictions))


# Train and evaluate the models on the test set
nb_classifier.fit(X_train, y_train)
nb_predictions = nb_classifier.predict(X_test)
print("Naive Bayes Test Accuracy: ", accuracy_score(y_test, nb_predictions))
print("Naive Bayes Test Precision: ", precision_score(y_test, nb_predictions))
print("Naive Bayes Test Recall: ", recall_score(y_test, nb_predictions))
print("Naive Bayes Test F1-Score: ", f1_score(y_test, nb_predictions))

# Train and evaluate the models on the test set
logreg.fit(X_train, y_train)
logreg_predictions = logreg.predict(X_test)
print("Logistic Regression Test Accuracy: ", accuracy_score(y_test, logreg_predictions))
print("Logistic Regression Test Precision: ", precision_score(y_test, logreg_predictions))
print("Logistic Regression Test Recall: ", recall_score(y_test, logreg_predictions))
print("Logistic Regression Test F1-Score: ", f1_score(y_test, logreg_predictions))

dt_classifier.fit(X_train, y_train)
dt_predictions = dt_classifier.predict(X_test)
print("Decision Tree Classifier Test Accuracy: ", accuracy_score(y_test, dt_predictions))
print("Decision Tree Classifier Test Precision: ", precision_score(y_test, dt_predictions))
print("Decision Tree Classifier Test Recall: ", recall_score(y_test, dt_predictions))
print("Decision Tree Classifier Test F1-Score: ", f1_score(y_test, dt_predictions))



#--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import Pipeline

df = pd.read_csv("F:\\python\\CSV\\spam.csv")
df.groupby('Category').describe()

df['spam'] = df['Category'].apply(lambda x: 1 if x=='spam' else 0)
X_train,X_test, y_train, y_test = train_test_split(df['Message'],df['spam'], test_size = 0.2)

cv = CountVectorizer()
X_trainset = cv.fit_transform(X_train.values)
X_trainset.toarray()[:3]

ml = MultinomialNB()
ml.fit(X_trainset, y_train)

emails = [
    "Hey mohan, can we get together to watch footbal game tomorrow?",
    "Upto 20% discount on parking, exclusive offer just for you. Dont miss this reward!"
]
emails_count = cv.transform(emails)
ml.predict(emails_count)

X_test_count = cv.transform(X_test)
print(ml.score(X_test_count, y_test))

sc = Pipeline([('vect', CountVectorizer()), ('nb', MultinomialNB())])
sc.fit(X_train, y_train)
print(sc.score(X_test,y_test))



#--------------------------------------------------------------------------------------------------------------------------------------------------


import speech_recognition as sr
import re

def transcribe_audio(audio_file):
    recognizer = sr.Recognizer()
    
    with sr.AudioFile(audio_file) as source:
        audio_data = recognizer.record(source)  # Record the entire audio file
        
    try:
        # Use Sphinx engine to transcribe the audio (offline)
        transcribed_text = recognizer.recognize_sphinx(audio_data)
        return transcribed_text
    except sr.UnknownValueError:
        print("Sphinx could not understand the audio")
    except sr.RequestError as e:
        print(f"Error occurred during Sphinx recognition; {e}")

    return None

def search_spoken_words(transcribed_text, search_word):
    # Search for the specified word in the transcribed text using regular expressions
    matches = re.findall(search_word, transcribed_text, re.IGNORECASE)
    return matches

# Example usage
if __name__ == "__main__":
    audio_file_path = "F:\dj.wav"  # Replace with the path to your audio file
    search_word = input("Enter the word to search for: ")

    transcribed_text = transcribe_audio(audio_file_path)
    if transcribed_text:
        found_matches = search_spoken_words(transcribed_text, search_word)
        if found_matches:
            print(f"Found {len(found_matches)} occurrences of '{search_word}' in the audio.")
            print("Occurrences found at the following indices:")
            for match in found_matches:
                print(match.start())
        else:
            print(f"No occurrences of '{search_word}' found in the audio.")
    else:
        print("Audio transcription failed.")




#--------------------------------------------------------------------------------------------------------------------------------------------------



import pandas as pd
import tensorflow
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
from keras.models import Model, load_model
from keras.layers import Embedding, LSTM, Dense, Input, Concatenate
import numpy as np
import pickle

# Load your dataset from an Excel file
df = pd.read_excel("F:\\python\\valid.xlsx")

# Convert numeric and alphanumeric columns to strings
df['context'] = df['context'].astype(str)
df['question'] = df['question'].astype(str)
df['answers'] = df['answers'].astype(str)

# Split the dataset into training and validation sets
train_percentage = 0.8
train_size = int(train_percentage * len(df))
train_df = df[:train_size]
valid_df = df[train_size:]

# Tokenize and pad sequences for context and question separately for training data
tokenizer = Tokenizer()
tokenizer.fit_on_texts(train_df['context'] + train_df['question'] + train_df['answers'])  # Include 'answers' column in tokenization

total_words = len(tokenizer.word_index) + 1

max_sequence_length = 50  # Choose an appropriate sequence length
train_context_sequences = tokenizer.texts_to_sequences(train_df['context'])
train_question_sequences = tokenizer.texts_to_sequences(train_df['question'])

train_context_padded = pad_sequences(train_context_sequences, maxlen=max_sequence_length, padding='post', truncating='post')
train_question_padded = pad_sequences(train_question_sequences, maxlen=max_sequence_length, padding='post', truncating='post')

# One-hot encode answer labels for training data
train_answer_sequences = tokenizer.texts_to_sequences(train_df['answers'])
max_answer_length = 10  # Choose an appropriate answer sequence length
train_answer_padded = pad_sequences(train_answer_sequences, maxlen=max_answer_length, padding='post', truncating='post')
num_classes = total_words  # Number of classes is equal to the total number of unique tokens

y_train = np.zeros((len(train_answer_padded), total_words), dtype=bool)

for i, sequence in enumerate(train_answer_padded):
    y_train[i, sequence] = 1

# Tokenize and pad sequences for context and question separately for validation data
valid_context_sequences = tokenizer.texts_to_sequences(valid_df['context'])
valid_question_sequences = tokenizer.texts_to_sequences(valid_df['question'])

valid_context_padded = pad_sequences(valid_context_sequences, maxlen=max_sequence_length, padding='post', truncating='post')
valid_question_padded = pad_sequences(valid_question_sequences, maxlen=max_sequence_length, padding='post', truncating='post')

# One-hot encode answer labels for validation data
valid_answer_sequences = tokenizer.texts_to_sequences(valid_df['answers'])
valid_answer_padded = pad_sequences(valid_answer_sequences, maxlen=max_answer_length, padding='post', truncating='post')

y_valid = np.zeros((len(valid_answer_padded), total_words), dtype=bool)

for i, sequence in enumerate(valid_answer_padded):
    y_valid[i, sequence] = 1

# Build the Keras model with multiple inputs
embedding_size = 100

# Context branch
context_input = Input(shape=(max_sequence_length,), name='context_input')
context_embedding = Embedding(total_words, embedding_size, input_length=max_sequence_length)(context_input)
context_lstm = LSTM(100)(context_embedding)

# Question branch
question_input = Input(shape=(max_sequence_length,), name='question_input')
question_embedding = Embedding(total_words, embedding_size, input_length=max_sequence_length)(question_input)
question_lstm = LSTM(100)(question_embedding)

# Concatenate both branches
merged = Concatenate()([context_lstm, question_lstm])

# Output layer
output = Dense(num_classes, activation='softmax')(merged)

# Create the model
model = Model(inputs=[context_input, question_input], outputs=output)

model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

# Train the model
model.fit([train_context_padded, train_question_padded], y_train, epochs=10, verbose=1, validation_data=([valid_context_padded, valid_question_padded], y_valid))

# Save the Keras model
model.save("model.keras")

# Save the tokenizer as pickle file
with open("tokenizer.pickle", "wb") as handle:
    pickle.dump(tokenizer, handle, protocol=pickle.HIGHEST_PROTOCOL)

# Load the Keras model
model = load_model("model.keras")

# Load the tokenizer from pickle file
with open("tokenizer.pickle", "rb") as handle:
    tokenizer = pickle.load(handle)

while True:
    input_context = [str(input("Context: "))]
    input_question = [str(input("Question: "))]

    input_context_sequences = tokenizer.texts_to_sequences(input_context)
    input_question_sequences = tokenizer.texts_to_sequences(input_question)

    max_sequence_length = 50
    input_context_padded = pad_sequences(input_context_sequences, maxlen=max_sequence_length, padding='post', truncating='post')
    input_question_padded = pad_sequences(input_question_sequences, maxlen=max_sequence_length, padding='post', truncating='post')

    # Generate predictions
    predicted_probabilities = model.predict([input_context_padded, input_question_padded], verbose=0)
    predicted_indices = np.argmax(predicted_probabilities, axis=1)
    predicted_answers = [tokenizer.index_word[idx] for idx in predicted_indices]

    print(f"Predicted answers: {predicted_answers}")




#--------------------------------------------------------------------------------------------------------------------------------------------------


import tensorflow as tf
from keras import layers, models
from keras.datasets import mnist
import matplotlib.pyplot as plt

# Load and preprocess the MNIST dataset
(train_images, train_labels), (test_images, test_labels) = mnist.load_data()
train_images, test_images = train_images / 255.0, test_images / 255.0

# Reshape the images to have a single channel (grayscale)
train_images = train_images.reshape((60000, 28, 28, 1))
test_images = test_images.reshape((10000, 28, 28, 1))

# Define the CNN model
model = models.Sequential()
model.add(layers.Conv2D(32, (3, 3), activation='relu', input_shape=(28, 28, 1)))
model.add(layers.MaxPooling2D((2, 2)))
model.add(layers.Conv2D(64, (3, 3), activation='relu'))
model.add(layers.MaxPooling2D((2, 2)))
model.add(layers.Conv2D(64, (3, 3), activation='relu'))

# Flatten the output for the fully connected layers
model.add(layers.Flatten())
model.add(layers.Dense(64, activation='relu'))
model.add(layers.Dense(10, activation='softmax'))

# Compile the model
model.compile(optimizer='adam',
              loss='sparse_categorical_crossentropy',
              metrics=['accuracy'])

# Display the model summary
model.summary()

# Train the model
model.fit(train_images, train_labels, epochs=1, validation_data=(test_images, test_labels))

# Evaluate the model on the test set
test_loss, test_acc = model.evaluate(test_images, test_labels)
print(f'Test accuracy: {test_acc}')

# Make predictions on some test images
predictions = model.predict(test_images[:5])

# Display the first few test images and their predicted labels
for i in range(5):
    plt.imshow(test_images[i].reshape(28, 28), cmap='gray')
    plt.title(f"True label: {test_labels[i]}, Predicted label: {tf.argmax(predictions[i])}")
    plt.show()




#--------------------------------------------------------------------------------------------------------------------------------------------------


from transformers import RobertaTokenizer, RobertaModel

# Load pre-trained RoBERTa tokenizer and model
tokenizer = RobertaTokenizer.from_pretrained("roberta-base")
model = RobertaModel.from_pretrained("roberta-base")

# Example text
text = "Hello, how are you doing today?"

# Tokenize input text
input_ids = tokenizer.encode(text, return_tensors="pt")

# Forward pass through the model
outputs = model(input_ids)

# Get the hidden states from the last layer
last_hidden_states = outputs.last_hidden_state

print("Input Text:", text)
print("Tokenized IDs:", input_ids)
print("Last Hidden States Shape:", last_hidden_states.shape)




#--------------------------------------------------------------------------------------------------------------------------------------------------


from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, classification_report
from sklearn.datasets import load_iris

# Load the Iris dataset
iris = load_iris()
X, y = iris.data, iris.target

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Standardize the features
scaler = StandardScaler()
X_train = scaler.fit_transform(X_train)
X_test = scaler.transform(X_test)

# Train a Random Forest Classifier
rf_classifier = RandomForestClassifier(n_estimators=100, random_state=42)
rf_classifier.fit(X_train, y_train)

# Make predictions on the test set
y_pred = rf_classifier.predict(X_test)

# Evaluate the model
accuracy = accuracy_score(y_test, y_pred)
classification_rep = classification_report(y_test, y_pred)

print("Machine Learning Model (Random Forest) Accuracy:", accuracy)
print("Classification Report:\n", classification_rep)




#--------------------------------------------------------------------------------------------------------------------------------------------------


from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
from tensorflow.keras.utils import to_categorical
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import accuracy_score, classification_report
from sklearn.datasets import load_iris

# Load the Iris dataset
iris = load_iris()
X, y = iris.data, iris.target

# Convert labels to one-hot encoding
y_one_hot = to_categorical(y)

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y_one_hot, test_size=0.2, random_state=42)

# Standardize the features
scaler = StandardScaler()
X_train = scaler.fit_transform(X_train)
X_test = scaler.transform(X_test)

# Build a simple neural network with Keras
model = Sequential()
model.add(Dense(8, input_dim=4, activation='relu'))
model.add(Dense(3, activation='softmax'))

# Compile the model
model.compile(loss='categorical_crossentropy', optimizer='adam', metrics=['accuracy'])

# Train the model
model.fit(X_train, y_train, epochs=50, batch_size=10, validation_data=(X_test, y_test))

# Evaluate the model
_, accuracy = model.evaluate(X_test, y_test)

print("Deep Learning Model (Neural Network) Accuracy:", accuracy)




#--------------------------------------------------------------------------------------------------------------------------------------------------


n = int(input('Enter the number: '))
s = [0,1]
for i in range(2,n+1):
    x = s[-1] + s[-2]
    s.append(x)
# int(num**0.5)**2 == num
print(s[-1])

def fibonacci_recursive(n):
    if n <= 1:
        return n
    else:
        return fibonacci_recursive(n - 1) + fibonacci_recursive(n - 2)

# Example Usage:
num_terms = 35
fibonacci_sequence = [fibonacci_recursive(i) for i in range(num_terms)]
print(fibonacci_sequence[-1])




#--------------------------------------------------------------------------------------------------------------------------------------------------



import tensorflow as tf

from keras import layers, models
import matplotlib.pyplot as plt
from keras.datasets import mnist
from keras.utils import to_categorical

(train_images, train_lables),(test_images,test_lables) = mnist.load_data()

train_images = train_images.reshape((60000, 28,28, 1)).astype('float32')/255
test_images = test_images.reshape((10000, 28,28, 1)).astype('float32')/255

train_lables = to_categorical(train_lables)
test_lables = to_categorical(test_lables)

model = models.Sequential()
model.add(layers.Conv2D(32, (3,3), activation='relu' , input_shape= (28,28,1) ))
model.add(layers.MaxPool2D((2,2)))
model.add(layers.Conv2D(64, (3,3), activation='relu'))
model.add(layers.MaxPool2D((2,2)))
model.add(layers.Conv2D(64, (3,3), activation='relu'))
model.add(layers.Flatten())
model.add(layers.Dense(64, activation='relu' ))
model.add(layers.Dense(10, activation='softmax' ))

model.compile(optimizer = 'adam', loss='categorical_crossentropy',metrics = ['accuracy'])

model.summary()

history = model.fit(train_images, train_lables, epochs = 5, batch_size = 64, validation_split = 0.2)

test_loss, test_acc = model.evaluate(test_images,test_lables)
print(f'Test accuracy: {test_acc}')

plt.plot(history.history['accuracy'], label='Training Accuracy')
plt.plot(history.history['val_accuracy'], label='Validation Accuracy')
plt.xlabel('Epoch')
plt.ylabel('Accuracy')
plt.legend()
plt.show()



#--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as pd
import json


with open("F:\\python\\Dataset\\dataset.json", 'r') as file:
    data = json.load(file).get('train', [])


data = {
    "train": [{
        "id": "56be85543aeaaa14008c9063", 
        "title": "Beyonc", 
        "context": "Beyonc Giselle Knowles-Carter is an American singer, songwriter, record producer and actress. Born and raised in Houston, Texas, she performed in various singing and dancing competitions as a child, and rose to fame in the late 1990s as lead singer of R&B girl-group Destiny's Child. Managed by her father, Mathew Knowles, the group became one of the world's best-selling girl groups of all time. Their hiatus saw the release of Beyonc's debut album, Dangerously in Love (2003), which established her as a solo artist worldwide, earned five Grammy Awards and featured the Billboard Hot 100 number-one singles", 
        "question": "When did Beyonce start becoming popular?", 
        "answers": {"text": ["in the late 1990s"], "answer_start": [269]}}]}

rows = []
for item in data['train']:
    title = item['title']
    context = item['context']
    question = item.get('question', '')
    answers_list = item.get('answers', {})
    
    if answers_list:
        answers = answers_list.get('text', [])
        answer_start = answers_list.get('answer_start', [])
    else:
        answers = []
        answer_start = []

    answers_str = ', '.join(map(str, answers)) if answers else ''
    answer_start_str = ', '.join(map(str, answer_start)) if answer_start else ''

    
    rows.append({
        'Title': title,
        'context': context,
        'Question': question,
        'Answers': answers_str,
        'Answer Start': answer_start_str,
    })

df = pd.DataFrame(rows)

print(df)
 
df.to_csv('RoBert.csv', index=False)




#--------------------------------------------------------------------------------------------------------------------------------------------------


import requests
import os

def download_repos(token, org_or_user, destination_folder="."):
    # Set the GitHub token
    headers = {"Authorization": f"token {token}"}

    # Initialize variables for pagination
    page = 1
    per_page = 100  # Adjust per_page based on your needs

    while True:
        # Make a request to the GitHub API to get repositories for the organization or user
        url = f"https://api.github.com/users/{org_or_user}/repos?page={page}&per_page={per_page}"
        response = requests.get(url, headers=headers)

        # Check if the request was successful
        if response.status_code == 200:
            repos = response.json()

            # Check if there are no more repositories
            if not repos:
                break

            # Loop through the repositories and clone each one
            for repo in repos:
                clone_url = repo["clone_url"]
                repo_name = repo["name"]
                repo_destination = os.path.join(destination_folder, repo_name)

                # Check if the repository already exists locally
                if os.path.exists(repo_destination):
                    print(f"Repository '{repo_name}' already exists. Skipping.")
                else:
                    # Clone the repository
                    os.system(f"git clone {clone_url} {repo_destination}")
                    print(f"Repository '{repo_name}' cloned successfully.")

            # Move to the next page
            page += 1
        else:
            print(f"Failed to retrieve repositories. Status code: {response.status_code}")
            break

if __name__ == "__main__":
    # Replace 'YOUR_GITHUB_TOKEN' with your GitHub personal access token
    github_token = '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'

    # Replace 'ORG_OR_USER' with the GitHub organization or user name
    organization_or_user = 'krishnaik06'

    # Specify the destination folder for cloning repositories
    destination_folder = 'F:\\python\\Git_Hub\\krish'

    # Create the destination folder if it doesn't exist
    os.makedirs(destination_folder, exist_ok=True)

    download_repos(github_token, organization_or_user, destination_folder)







#--------------------------------------------------------------------------------------------------------------------------------------------------


import requests
import os

def download_repos(token, org_or_user, target_folder):
    # Set the GitHub token
    headers = {"Authorization": f"token {token}"}

    # Get the list of repositories for the organization or user
    response = requests.get(f"https://api.github.com/users/{org_or_user}/repos", headers=headers)

    # Check if the request was successful
    if response.status_code == 200:
        repos = response.json()

        # Loop through the repositories and clone each one
        for repo in repos:
            clone_url = repo["clone_url"]
            repo_name = repo["name"]
            target_path = os.path.join(target_folder, repo_name)

            # Check if the repository already exists locally
            if os.path.exists(target_path):
                print(f"Repository '{repo_name}' already exists in '{target_folder}'. Skipping.")
            else:
                # Clone the repository into the specified target folder
                os.system(f"git clone {clone_url} {target_path}")
                print(f"Repository '{repo_name}' cloned successfully into '{target_folder}'.")
    else:
        print(f"Failed to retrieve repositories. Status code: {response.status_code}")

if __name__ == "__main__":
    # Replace 'YOUR_GITHUB_TOKEN' with your GitHub personal access token
    github_token = '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'

    # Replace 'ORG_OR_USER' with the GitHub organization or user name
    organization_or_user = 'suneelbvs'

    # Specify the target folder where repositories will be cloned
    target_folder = 'F:\\python\\Git_Hub\\suneelbvs'

    download_repos(github_token, organization_or_user, target_folder)




#--------------------------------------------------------------------------------------------------------------------------------------------------



# export GIT_COMMITTER_DATE='Wed Dec 21 11:51:39 IST 2022'
# git commit --amend --no-edit --date='Wed Dec 21 11:51:39 IST 2022'
# unset GIT_COMMITTER_DATE


#--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as pd
from sklearn.model_selection import train_test_split
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
from keras.models import Sequential
from keras.layers import Embedding, LSTM, Dense

# Load your dataset (assuming it's a CSV file)
# Replace 'your_dataset.csv' with the actual file path
dataset = pd.read_csv("F:\\python\\CSV\\testtt.csv")

# Assuming 'text' is the input sequence and 'ending' is the target sequence
input_texts = dataset['text'].astype(str).tolist()
target_texts = dataset['ended'].astype(str).tolist()

# Tokenize input and target sequences
tokenizer_input = Tokenizer()
tokenizer_input.fit_on_texts(input_texts)
input_sequences = tokenizer_input.texts_to_sequences(input_texts)

tokenizer_target = Tokenizer()
tokenizer_target.fit_on_texts(target_texts)
target_sequences = tokenizer_target.texts_to_sequences(target_texts)

# Pad sequences to a fixed length
max_sequence_length = max(max(map(len, input_sequences)), max(map(len, target_sequences)))
input_sequences_padded = pad_sequences(input_sequences, maxlen=max_sequence_length, padding='post')
target_sequences_padded = pad_sequences(target_sequences, maxlen=max_sequence_length, padding='post')

# Define the model
vocab_size_input = len(tokenizer_input.word_index) + 1
vocab_size_target = len(tokenizer_target.word_index) + 1

embedding_dim = 50  # Adjust based on your dataset and computational resources

model = Sequential()
model.add(Embedding(input_dim=vocab_size_input, output_dim=embedding_dim, input_length=max_sequence_length))
model.add(LSTM(50))
model.add(Dense(vocab_size_target, activation='softmax'))

model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy'])

# Split the data into training and validation sets
input_train, input_val, target_train, target_val = train_test_split(
    input_sequences_padded, target_sequences_padded, test_size=0.2, random_state=42
)

# Train the model
model.fit(input_train, target_train, epochs=10, validation_data=(input_val, target_val))

# Now, you can use the trained model for sequence generation.
# You may need to adapt the code based on the specifics of your dataset and task.
# Assume 'model' is your trained Keras model
# Assume 'tokenizer_input' and 'tokenizer_target' are your trained tokenizers

def generate_sequence(input_text):
    # Tokenize the input text
    input_sequence = tokenizer_input.texts_to_sequences([input_text])
    
    # Pad the input sequence
    input_sequence_padded = pad_sequences(input_sequence, maxlen=max_sequence_length, padding='post')
    
    # Use the model to generate the output sequence
    output_sequence_padded = model.predict(input_sequence_padded)[0]
    
    # Convert the output sequence back to text
    output_sequence = [tokenizer_target.index_word[idx] for idx in output_sequence_padded.argmax(axis=-1)]
    
    # Join the tokens to get the final output text
    output_text = ' '.join(output_sequence)
    
    return output_text

# Example usage
input_text = "Your input text goes here."
generated_output = generate_sequence(input_text)

print("Input Text: ", input_text)
print("Generated Output: ", generated_output)




#--------------------------------------------------------------------------------------------------------------------------------------------------



import os
import easyocr
import PyPDF2
import pandas as pd
from PIL import Image

# Function to read text from an image using EasyOCR
def read_text_from_image(file_path):
    reader = easyocr.Reader(['en'])  # Language selection, 'en' for English
    with Image.open(file_path) as img:
        text = reader.readtext(img)
    return '\n'.join([result[1] for result in text])
file_path = "F:\\ghg.png"
read_text_from_image(file_path)

# import sys
# print(sys.executable)



#--------------------------------------------------------------------------------------------------------------------------------------------------


import cv2

# Open a video capture object (0 is usually the default camera)
cap = cv2.VideoCapture(0)

# Check if the camera opened successfully
if not cap.isOpened():
    print("Error: Could not open camera.")
    exit()

# Define the window name
window_name = 'Video Input'

# Create a window to display the video
cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)

while True:
    # Read a frame from the camera
    ret, frame = cap.read()

    # Check if the frame was successfully captured
    if not ret:
        print("Error: Could not read frame.")
        break

    # Display the frame in the window
    cv2.imshow(window_name, frame)

    # Break the loop if the 'q' key is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the video capture object and close the window
cap.release()
cv2.destroyAllWindows()




#--------------------------------------------------------------------------------------------------------------------------------------------------



import cv2
import numpy as np

# Open a video capture object (0 is usually the default camera)
cap = cv2.VideoCapture(0)

# Check if the camera opened successfully
if not cap.isOpened():
    print("Error: Could not open camera.")
    exit()

# Define the window name
window_name = 'Hazard Detection'

# Create a window to display the video
cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)

while True:
    # Read a frame from the camera
    ret, frame = cap.read()

    # Check if the frame was successfully captured
    if not ret:
        print("Error: Could not read frame.")
        break

    # Convert the frame to HSV color space
    hsv_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

    # Define a color range for yellow (you may need to adjust these values)
    lower_yellow = np.array([20, 100, 100])
    upper_yellow = np.array([30, 255, 255])

    # Threshold the frame to get only yellow regions
    yellow_mask = cv2.inRange(hsv_frame, lower_yellow, upper_yellow)

    # Bitwise-AND the original frame and the mask
    result_frame = cv2.bitwise_and(frame, frame, mask=yellow_mask)

    # Display the original frame and the result side by side
    display_frame = np.hstack((frame, result_frame))
    cv2.imshow(window_name, display_frame)

    # Break the loop if the 'q' key is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the video capture object and close the window
cap.release()
cv2.destroyAllWindows()





#--------------------------------------------------------------------------------------------------------------------------------------------------


import cv2
import numpy as np
import tensorflow as tf

# Load the pre-trained model
model_path = "path/to/your/model"  # Replace with the actual path to your model
model = tf.saved_model.load(model_path)

# Open a video capture object (0 is usually the default camera)
cap = cv2.VideoCapture(0)

# Check if the camera opened successfully
if not cap.isOpened():
    print("Error: Could not open camera.")
    exit()

# Define the window name
window_name = 'Helmet Detection'

# Create a window to display the video
cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)

while True:
    # Read a frame from the camera
    ret, frame = cap.read()

    # Check if the frame was successfully captured
    if not ret:
        print("Error: Could not read frame.")
        break

    # Preprocess the frame for the model
    input_tensor = tf.convert_to_tensor(frame)
    input_tensor = input_tensor[tf.newaxis, ...]

    # Make predictions
    detections = model(input_tensor)

    # Extract bounding boxes, classes, and scores from the detections
    boxes = detections['detection_boxes'][0].numpy()
    classes = detections['detection_classes'][0].numpy().astype(np.int32)
    scores = detections['detection_scores'][0].numpy()

    # Draw bounding boxes on the frame
    for i in range(len(scores)):
        if scores[i] > 0.5 and classes[i] == 1:  # Class 1 corresponds to person in COCO dataset
            ymin, xmin, ymax, xmax = boxes[i]
            ymin, xmin, ymax, xmax = int(ymin * frame.shape[0]), int(xmin * frame.shape[1]), \
                                     int(ymax * frame.shape[0]), int(xmax * frame.shape[1])
            cv2.rectangle(frame, (xmin, ymin), (xmax, ymax), (0, 255, 0), 2)

    # Display the frame with bounding boxes
    cv2.imshow(window_name, frame)

    # Break the loop if the 'q' key is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the video capture object and close the window
cap.release()
cv2.destroyAllWindows()




#--------------------------------------------------------------------------------------------------------------------------------------------------



import cv2
import numpy as np

# Open a video capture object (0 is usually the default camera)
cap = cv2.VideoCapture(0)

# Check if the camera opened successfully
if not cap.isOpened():
    print("Error: Could not open camera.")
    exit()

# Define the window name
window_name = 'Helmet Detection'

# Create a window to display the video
cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)

# Define the color range for helmets (you may need to adjust these values)
lower_helmet_color = np.array([0, 0, 100])
upper_helmet_color = np.array([50, 50, 255])

while True:
    # Read a frame from the camera
    ret, frame = cap.read()

    # Check if the frame was successfully captured
    if not ret:
        print("Error: Could not read frame.")
        break

    # Convert the frame to HSV color space
    hsv_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

    # Threshold the frame to get only regions with helmet color
    helmet_mask = cv2.inRange(hsv_frame, lower_helmet_color, upper_helmet_color)

    # Find contours in the mask
    contours, _ = cv2.findContours(helmet_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # Draw bounding boxes around detected helmets
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

    # Display the frame with bounding boxes
    cv2.imshow(window_name, frame)

    # Break the loop if the 'q' key is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the video capture object and close the window
cap.release()
cv2.destroyAllWindows()



#--------------------------------------------------------------------------------------------------------------------------------------------------



import jsonlines
import tensorflow as tf
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences


file_path = "F:\\python\\data\\web.jsonl"

with jsonlines.open(file_path, mode ='r') as reader:
        data = []
        for line in reader:
            data.append(line)

# Extract text
text = [example["text"] for example in data]

# Tokenize text
tokenizer = Tokenizer()
tokenizer.fit_on_texts(text)

# Convert text to sequences
sequences = tokenizer.texts_to_sequences(text)

# Create input sequences and target sequences
input_sequences = [seq[:-1] for seq in sequences]
target_sequences = [seq[1:] for seq in sequences]

# Pad sequences for fixed input length
max_length = max(map(len, sequences))
input_padded = pad_sequences(input_sequences, maxlen=max_length, padding='post')
target_padded = pad_sequences(target_sequences, maxlen=max_length, padding='post')

# Convert target sequences to categorical
target_categorical = tf.keras.utils.to_categorical(target_padded, num_classes=len(tokenizer.word_index) + 1)

# Define the model
model = tf.keras.Sequential([
    tf.keras.layers.Embedding(input_dim=len(tokenizer.word_index) + 1, output_dim=100, input_length=max_length),
    tf.keras.layers.LSTM(100, return_sequences=True),
    tf.keras.layers.Dense(len(tokenizer.word_index) + 1, activation='softmax')
])

# Compile the model
model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

# Train the model
model.fit(input_padded, target_categorical, epochs=10)


# Use the trained model for inference
test_prompt = "House Majority Whip Steve Scalise has been discharged"
test_prompt_sequence = tokenizer.texts_to_sequences([test_prompt])
test_prompt_padded = pad_sequences(test_prompt_sequence, maxlen=max_length, padding='post')

# Initial prediction
predicted_sequence = model.predict(test_prompt_padded)

# Generate the next word iteratively
for _ in range(5):  # You can adjust the number of words to generate
    next_word_index = tf.argmax(predicted_sequence[0, -1, :]).numpy()
    next_word = tokenizer.index_word[next_word_index]
    test_prompt += " " + next_word

    # Update the input sequence for the next prediction
    test_prompt_sequence = tokenizer.texts_to_sequences([test_prompt])
    test_prompt_padded = pad_sequences(test_prompt_sequence, maxlen=max_length, padding='post')

    # Predict the next word
    predicted_sequence = model.predict(test_prompt_padded)

print("Generated Text:", test_prompt)





#--------------------------------------------------------------------------------------------------------------------------------------------------



import pandas as pd
from sklearn.model_selection import train_test_split
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
from keras.models import Sequential
from keras.layers import Embedding, LSTM, Dense

# Load your dataset (assuming it's a CSV file)
# Replace 'your_greetings_dataset.csv' with the actual file path
dataset = pd.read_csv("F:\\python\\CSV\\your_greetings_dataset.csv")

print(dataset.columns)

# Assuming 'input_text' is the input sequence and 'response' is the target sequence
input_texts = dataset['input_text'].astype(str).tolist()
target_texts = dataset['response'].astype(str).tolist()

# Tokenize input and target sequences
tokenizer_input = Tokenizer()
tokenizer_input.fit_on_texts(input_texts)
input_sequences = tokenizer_input.texts_to_sequences(input_texts)

tokenizer_target = Tokenizer()
tokenizer_target.fit_on_texts(target_texts)
target_sequences = tokenizer_target.texts_to_sequences(target_texts)

# Pad sequences to a fixed length
max_sequence_length = max(max(map(len, input_sequences)), max(map(len, target_sequences)))
input_sequences_padded = pad_sequences(input_sequences, maxlen=max_sequence_length, padding='post')
target_sequences_padded = pad_sequences(target_sequences, maxlen=max_sequence_length, padding='post')

# Define the model
vocab_size_input = len(tokenizer_input.word_index) + 1
vocab_size_target = len(tokenizer_target.word_index) + 1

embedding_dim = 50  # Adjust based on your dataset and computational resources

model = Sequential()
model.add(Embedding(input_dim=vocab_size_input, output_dim=embedding_dim, input_length=max_sequence_length))
model.add(LSTM(50))
model.add(Dense(vocab_size_target, activation='softmax'))

model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy'])

# Split the data into training and validation sets
input_train, input_val, target_train, target_val = train_test_split(
    input_sequences_padded, target_sequences_padded, test_size=0.2, random_state=42
)

# Train the model
model.fit(input_train, target_train, epochs=10, validation_data=(input_val, target_val))

# Now, you can use the trained model for sequence generation.
# Assume 'model' is your trained Keras model
# Assume 'tokenizer_input' and 'tokenizer_target' are your trained tokenizers

def generate_response(input_text):
    # Tokenize the input text
    input_sequence = tokenizer_input.texts_to_sequences([input_text])
    
    # Pad the input sequence
    input_sequence_padded = pad_sequences(input_sequence, maxlen=max_sequence_length, padding='post')
    
    # Use the model to generate the output sequence
    output_sequence_padded = model.predict(input_sequence_padded)[0]
    
    # Convert the output sequence back to text
    output_sequence = [tokenizer_target.index_word[idx] for idx in output_sequence_padded.argmax(axis=-1)]
    
    # Join the tokens to get the final output text
    output_text = ' '.join(output_sequence)
    
    return output_text

# Example usage
input_text = "hi, how are you"
generated_response = generate_response(input_text)

print("Input Text: ", input_text)
print("Generated Response: ", generated_response)




#--------------------------------------------------------------------------------------------------------------------------------------------------



import jsonlines
import pandas as pd

file_path = "F:\\python\\data\\web.jsonl"

data = []

with jsonlines.open(file_path, mode='r') as reader:
    for line in reader:
        data.append({
            'id': line['id'],
            'ended': line['ended'],
            'length': line['length'],
            'text': line['text']
        })

df = pd.DataFrame(data)
print(df)
df.to_csv('test.csv')


#--------------------------------------------------------------------------------------------------------------------------------------------------


# from transformers import BertTokenizer, BertForQuestionAnswering
# import torch
# import pandas as pd

# df = pd.read_csv("F:\\python\\CSV\\single_qna.csv")

# # Load pre-trained BERT model and tokenizer
# tokenizer = BertTokenizer.from_pretrained("bert-base-uncased")
# model = BertForQuestionAnswering.from_pretrained("bert-base-uncased")

# # Sample context and question
# context = df['Answer']
# question = df['Question']

# # Tokenize input
# inputs = tokenizer(context, question, return_tensors="pt")

# # Get the answer
# outputs = model(**inputs)
# answer_start = torch.argmax(outputs.start_logits)
# answer_end = torch.argmax(outputs.end_logits) + 1
# answer = tokenizer.convert_tokens_to_string(tokenizer.convert_ids_to_tokens(inputs["input_ids"][0][answer_start:answer_end]))

# print("Question:", question)
# print("Answer:", answer)

from transformers import BertTokenizer, BertForQuestionAnswering
import torch
import pandas as pd

df = pd.read_csv("F:\\python\\CSV\\single_qna.csv")

# Load pre-trained BERT model and tokenizer
tokenizer = BertTokenizer.from_pretrained("bert-base-uncased")
model = BertForQuestionAnswering.from_pretrained("bert-base-uncased")

# Iterate over each row in the DataFrame
for index, row in df.iterrows():
    context = row['Answer']
    question = row['Question']

    # Tokenize input
    inputs = tokenizer(context, question, return_tensors="pt")

    # Get the answer
    outputs = model(**inputs)
    answer_start = torch.argmax(outputs.start_logits)
    answer_end = torch.argmax(outputs.end_logits) + 1
    answer = tokenizer.convert_tokens_to_string(
        tokenizer.convert_ids_to_tokens(inputs["input_ids"][0][answer_start:answer_end])
    )

    print(f"Question ({index + 1}):", question)
    print(f"Answer ({index + 1}):", answer)
    print()





#--------------------------------------------------------------------------------------------------------------------------------------------------


import cv2
from matplotlib import pyplot as plt

# Read an image from file
image_path = 'F:\ghg.png'
image = cv2.imread(image_path)

# Check if the image was successfully loaded
if image is not None:
    # Display the image using Matplotlib
    plt.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
    plt.axis('off')  # Turn off axis labels
    plt.show()
else:
    print(f"Unable to read the image at {image_path}")




#--------------------------------------------------------------------------------------------------------------------------------------------------



# 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
    
from langchain import HuggingFaceHub
import os
os.environ["HUGGINGFACEHUB_API_TOKEN"]= "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"
llm = HuggingFaceHub(repo_id='google/flan-t5-large', model_kwargs={'temperature': 0, 'max_length':64})
text = 'what is the capital of australia'
print(llm.predict(text))
from langchain.prompts import PromptTemplate

promt_template = PromptTemplate(input_variables=['country'], template="what is th capital of {country}")

promt_template.format(country = 'India')
from langchain.chains import LLMChain

chain = LLMChain(llm = llm, prompt = promt_template)
chain.run('India')

#--------------------------------------------------------------------------------------------------------------------------------------------------


import cv2
import tensorflow as tf
import tensorflow_hub as hub
import numpy as np

model_url = "https://tfhub.dev/google/tf2-preview/mobilenet_v2/classification/4"
model = tf.keras.Sequential([hub.KerasLayer(model_url)])

image = "F:\\path\\MG_0411-web.jpg"
image = cv2.imread(image)
image_con = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
image = image_con/255

input_size=(224,224)
image_res = cv2.resize(image, input_size)

img_exp = np.expand_dims(image_res, axis=0)

predictions = model.predict(img_exp)

threshold = 0.4
predict_class = 1 if predictions[0][0] > threshold else 0

print(predictions[0][0])

if predict_class == 1:
    print('Helmet detected')
else:
    print('helmet not detected')



#--------------------------------------------------------------------------------------------------------------------------------------------------



import tensorflow as tf
from keras import layers, models
from keras.preprocessing.image import ImageDataGenerator

# Define the CNN model
model = models.Sequential()
model.add(layers.Conv2D(32, (3, 3), activation='relu', input_shape=(224, 224, 3)))
model.add(layers.MaxPooling2D((2, 2)))
model.add(layers.Conv2D(64, (3, 3), activation='relu'))
model.add(layers.MaxPooling2D((2, 2)))
model.add(layers.Conv2D(128, (3, 3), activation='relu'))
model.add(layers.MaxPooling2D((2, 2)))
model.add(layers.Flatten())
model.add(layers.Dense(128, activation='relu'))
model.add(layers.Dense(1, activation='sigmoid'))

# Compile the model
model.compile(optimizer='adam',
              loss='binary_crossentropy',
              metrics=['accuracy'])

# Define paths to your training and validation datasets
train_data_dir = 'path/to/your/training_data'
val_data_dir = 'path/to/your/validation_data'

# Set up data generators with preprocessing
train_datagen = ImageDataGenerator(rescale=1./255,
                                   shear_range=0.2,
                                   zoom_range=0.2,
                                   horizontal_flip=True)

val_datagen = ImageDataGenerator(rescale=1./255)

# Define batch size and image dimensions
batch_size = 32
img_height, img_width = 224, 224

# Create data generators
train_generator = train_datagen.flow_from_directory(
    train_data_dir,
    target_size=(img_height, img_width),
    batch_size=batch_size,
    class_mode='binary'  # Set to 'binary' for binary classification (helmet or not)
)

val_generator = val_datagen.flow_from_directory(
    val_data_dir,
    target_size=(img_height, img_width),
    batch_size=batch_size,
    class_mode='binary'
)

# Train the model
epochs = 10
model.fit(train_generator, epochs=epochs, validation_data=val_generator)

# Save the trained model
model.save('helmet_detection_model.h5')



#--------------------------------------------------------------------------------------------------------------------------------------------------



from roboflow import Roboflow
import pandas as pd
rf = Roboflow(api_key="8NrDQJZHRWY5KIIIbxGr")
project = rf.workspace().project("bike-helmet-detection-2vdjo")
model = project.version(2).model

# infer on a local image
pred = model.predict("F:\\path\\wo.jpg", confidence=40, overlap=30).json()
print(pred)
# df = pd.DataFrame(pred)
# print(df)

# Given output
output = pred

# Extracting relevant information from the output
prediction_data = output['predictions'][0]
image_data = output['image']

# Creating a DataFrame
df = pd.DataFrame([prediction_data])

# Adding image information as columns
df['image_width'] = image_data['width']
df['image_height'] = image_data['height']

# Reordering columns for better readability
df = df[['x', 'y', 'width', 'height', 'confidence', 'class', 'class_id', 'image_width', 'image_height', 'prediction_type']]

# Displaying the DataFrame
print(df)


# visualize your prediction
# model.predict("your_image.jpg", confidence=40, overlap=30).save("prediction.jpg")


#--------------------------------------------------------------------------------------------------------------------------------------------------



import os
import pandas as pd
from roboflow import Roboflow

# Initialize Roboflow client
rf = Roboflow(api_key="8NrDQJZHRWY5KIIIbxGr")

# Specify project and model
project = rf.workspace().project("bike-helmet-detection-2vdjo")
model_version = 2
model = project.version(model_version).model

# Path to the folder containing images
folder_path = "C:\\Users\\bbhar\\Downloads\\Compressed\\archive\\images"

# List all files in the folder
image_files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

# Create an empty DataFrame to store results
df_result = pd.DataFrame()

# Make predictions for each image
for image_file in image_files:
    # Construct the full path to the image
    image_path = os.path.join(folder_path, image_file)
    
    # Make predictions for the image
    pred = model.predict(image_path, confidence=40, overlap=30).json()
    
    # Check if 'predictions' list is not empty
    if 'predictions' in pred and pred['predictions']:
        # Extracting relevant information from the output
        prediction_data = pred['predictions'][0]
        image_data = pred['image']

        # Creating a DataFrame for the current image
        df = pd.DataFrame([prediction_data])

        # Adding image information as columns
        df['image_width'] = image_data['width']
        df['image_height'] = image_data['height']

        # Reordering columns for better readability
        df = df[['x', 'y', 'width', 'height', 'confidence', 'class', 'class_id', 'image_width', 'image_height', 'prediction_type']]

        # Append the DataFrame for the current image to the overall result DataFrame
        df_result = pd.concat([df_result, df], ignore_index=True)
    else:
        print(f"No predictions for {image_file}")

# Displaying the final DataFrame with results
print(df_result)
df_result.to_csv('F:\\python\\helmet.csv' )





#--------------------------------------------------------------------------------------------------------------------------------------------------




import os
import cv2
import xml.etree.ElementTree as ET
import numpy as np
from keras.preprocessing.image import ImageDataGenerator
import tensorflow as tf
from tensorflow import keras

# Define the CNN model
model = keras.models.Sequential()
model.add(keras.layers.Conv2D(32, (3, 3), activation='relu', input_shape=(224, 224, 3)))
model.add(keras.layers.MaxPooling2D((2, 2)))
model.add(keras.layers.Conv2D(64, (3, 3), activation='relu'))
model.add(keras.layers.MaxPooling2D((2, 2)))
model.add(keras.layers.Conv2D(128, (3, 3), activation='relu'))
model.add(keras.layers.MaxPooling2D((2, 2)))
model.add(keras.layers.Flatten())
model.add(keras.layers.Dense(128, activation='relu'))
model.add(keras.layers.Dense(1, activation='sigmoid'))

# Compile the model
model.compile(optimizer='adam',
              loss='binary_crossentropy',
              metrics=['accuracy'])

# Define paths to your training and validation datasets
train_images_dir = "C:\\Users\\bbhar\\Downloads\\Compressed\\archive_2\\train\\images"
train_annotations_dir = "C:\\Users\\bbhar\\Downloads\\Compressed\\archive_2\\train\\annotations"

val_images_dir = "C:\\Users\\bbhar\\Downloads\\Compressed\\archive_2\\test\\img"
val_annotations_dir = "C:\\Users\\bbhar\\Downloads\\Compressed\\archive_2\\test\\anot"

# Set up data generators with preprocessing
train_datagen = ImageDataGenerator(rescale=1./255,
                                   shear_range=0.2,
                                   zoom_range=0.2,
                                   horizontal_flip=True)

val_datagen = ImageDataGenerator(rescale=1./255)

# Define batch size and image dimensions
batch_size = 32
img_height, img_width = 224, 224

# Create data generators
train_generator = train_datagen.flow_from_directory(
    train_images_dir,
    target_size=(img_height, img_width),
    batch_size=batch_size,
    class_mode='binary'  # Set to 'binary' for binary classification (helmet or not)
)

val_generator = val_datagen.flow_from_directory(
    val_images_dir,
    target_size=(img_height, img_width),
    batch_size=batch_size,
    class_mode='binary'
)

# Function to parse XML annotations and extract bounding box information
def parse_annotation(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    boxes = []
    labels = []

    for obj in root.findall('object'):
        name = obj.find('name').text
        xmin = int(obj.find('bndbox/xmin').text)
        ymin = int(obj.find('bndbox/ymin').text)
        xmax = int(obj.find('bndbox/xmax').text)
        ymax = int(obj.find('bndbox/ymax').text)

        boxes.append([xmin, ymin, xmax, ymax])
        labels.append(name)

    return boxes, labels

# Function to load images and annotations from the generator
def load_data_from_generator(generator, images_dir, annotations_dir):
    while True:
        batch_x, batch_y = next(generator)

        images = []
        annotations = []

        for i in range(len(batch_x)):
            image_path = batch_x[i]
            annotation_path = os.path.join(annotations_dir, os.path.splitext(os.path.basename(image_path))[0] + '.xml')

            image = cv2.imread(image_path)
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            image = image / 255.0

            boxes, labels = parse_annotation(annotation_path)

            images.append(image)
            annotations.append({'boxes': boxes, 'labels': labels})

        yield np.array(images), np.array(annotations)

# Set the number of epochs
epochs = 10  # You should define the number of epochs here

# Example of how to use the data generator with images and annotations
for epoch in range(epochs):
    steps_per_epoch = len(train_generator)  # Number of batches per epoch
    validation_steps = len(val_generator)   # Number of batches for validation

    model.fit(load_data_from_generator(train_generator, train_images_dir, train_annotations_dir),
              epochs=1,
              steps_per_epoch=steps_per_epoch,
              validation_data=load_data_from_generator(val_generator, val_images_dir, val_annotations_dir),
              validation_steps=validation_steps)

# Save the trained model
model.save('helmet_detection_model.h5')




#--------------------------------------------------------------------------------------------------------------------------------------------------


import pandas as pd
import seaborn as sn
from sklearn.datasets import load_iris
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, confusion_matrix
import pickle

df = sn.load_dataset('iris')
print(df.target_names)
y =df['species']
x=df.drop('species',axis=1)
model = RandomForestClassifier
model.fit(x,y)
model.predict(x)

with open('Randomforest.pkl','wb') as model_pkl:
    pickle.dump(model, model_pkl)




#--------------------------------------------------------------------------------------------------------------------------------------------------


import cv2
import os

def extract_frames(video_path, output_folder):
    # Open the video file
    cap = cv2.VideoCapture(video_path)

    # Create the output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Get video information
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps = cap.get(cv2.CAP_PROP_FPS)

    print(f"Total Frames: {frame_count}")
    print(f"Frames Per Second (FPS): {fps}")

    # Loop through frames and save as images
    for frame_number in range(frame_count):
        # Read the frame
        ret, frame = cap.read()

        # Break the loop if the video is over
        if not ret:
            break

        # Save the frame as an image
        frame_filename = os.path.join(output_folder, f"frame_{frame_number:04d}.jpg")
        cv2.imwrite(frame_filename, frame)

    # Release the video capture object
    cap.release()

if __name__ == "__main__":
    # Specify the path to the video file
    video_path = "path/to/your/video.mp4"

    # Specify the output folder for frames
    output_folder = "output_frames"

    # Extract frames from the video
    extract_frames(video_path, output_folder)




#--------------------------------------------------------------------------------------------------------------------------------------------------



from __future__ import print_function
import keras
from keras.preprocessing.image import ImageDataGenerator
from keras.models import Sequential
from keras.layers import Dense,Dropout,Activation,Flatten,BatchNormalization
from keras.layers import Conv2D,MaxPooling2D
import os

num_classes = 5
img_rows,img_cols = 48,48
batch_size = 32

train_data_dir = r"F:\python\data\archive\train"
validation_data_dir = r"F:\python\data\archive\test"
train_datagen = ImageDataGenerator(
					rescale=1./255,
					rotation_range=30,
					shear_range=0.3,
					zoom_range=0.3,
					width_shift_range=0.4,
					height_shift_range=0.4,
					horizontal_flip=True,
					fill_mode='nearest')

validation_datagen = ImageDataGenerator(rescale=1./255)

train_generator = train_datagen.flow_from_directory(
					train_data_dir,
					color_mode='grayscale',
					target_size=(img_rows,img_cols),
					batch_size=batch_size,
					class_mode='categorical',
					shuffle=True)

validation_generator = validation_datagen.flow_from_directory(
							validation_data_dir,
							color_mode='grayscale',
							target_size=(img_rows,img_cols),
							batch_size=batch_size,
							class_mode='categorical',
							shuffle=True)


model = Sequential()

# Block-1


model.add(Conv2D(32,(3,3),padding='same',kernel_initializer='he_normal',input_shape=(img_rows,img_cols,1)))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Conv2D(32,(3,3),padding='same',kernel_initializer='he_normal',input_shape=(img_rows,img_cols,1)))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(MaxPooling2D(pool_size=(2,2)))
model.add(Dropout(0.2))


# Block-2 

model.add(Conv2D(64,(3,3),padding='same',kernel_initializer='he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Conv2D(64,(3,3),padding='same',kernel_initializer='he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(MaxPooling2D(pool_size=(2,2)))
model.add(Dropout(0.2))

# Block-3

model.add(Conv2D(128,(3,3),padding='same',kernel_initializer='he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Conv2D(128,(3,3),padding='same',kernel_initializer='he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(MaxPooling2D(pool_size=(2,2)))
model.add(Dropout(0.2))

# Block-4 

model.add(Conv2D(256,(3,3),padding='same',kernel_initializer='he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Conv2D(256,(3,3),padding='same',kernel_initializer='he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(MaxPooling2D(pool_size=(2,2)))
model.add(Dropout(0.2))

# Block-5

model.add(Flatten())
model.add(Dense(64,kernel_initializer='he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Dropout(0.5))

# Block-6

model.add(Dense(64,kernel_initializer='he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Dropout(0.5))

# Block-7

model.add(Dense(num_classes,kernel_initializer='he_normal'))
model.add(Activation('softmax'))

print(model.summary())

from keras.optimizers import RMSprop,SGD,Adam
from keras.callbacks import ModelCheckpoint, EarlyStopping, ReduceLROnPlateau

checkpoint = ModelCheckpoint('Emotion_little_vgg.h5',
                             monitor='val_loss',
                             mode='min',
                             save_best_only=True,
                             verbose=1)

earlystop = EarlyStopping(monitor='val_loss',
                          min_delta=0,
                          patience=3,
                          verbose=1,
                          restore_best_weights=True
                          )

reduce_lr = ReduceLROnPlateau(monitor='val_loss',
                              factor=0.2,
                              patience=3,
                              verbose=1,
                              min_delta=0.0001)

callbacks = [earlystop,checkpoint,reduce_lr]

model.compile(loss='categorical_crossentropy',
              optimizer = Adam(lr=0.001),
              metrics=['accuracy'])

nb_train_samples = 24176
nb_validation_samples = 3006
epochs=25

history=model.fit_generator(
                train_generator,
                steps_per_epoch=nb_train_samples//batch_size,
                epochs=epochs,
                callbacks=callbacks,
                validation_data=validation_generator,
                validation_steps=nb_validation_samples//batch_size)



from keras.models import load_model
from time import sleep
from keras.preprocessing.image import img_to_array
from keras.preprocessing import image
import cv2
import numpy as np

face_classifier = cv2.CascadeClassifier(r"F:\python\data\Facial-Expressions-Recognition-master\haarcascade_frontalface_default.xml")
classifier =load_model(r"F:\python\Emotion_vgg.h5")

class_labels = ['Angry','Happy','Neutral','Sad','Surprise']

# def face_detector(img):
#     # Convert image to grayscale
#     gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
#     faces = face_classifier.detectMultiScale(gray,1.3,5)
#     if faces is ():
#         return (0,0,0,0),np.zeros((48,48),np.uint8),img

#     for (x,y,w,h) in faces:
#         cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)
#         roi_gray = gray[y:y+h,x:x+w]

#     try:
#         roi_gray = cv2.resize(roi_gray,(48,48),interpolation=cv2.INTER_AREA)
#     except:
#         return (x,w,y,h),np.zeros((48,48),np.uint8),img
#     return (x,w,y,h),roi_gray,img


cap = cv2.VideoCapture(0)



while True:
    # Grab a single frame of video
    ret, frame = cap.read()
    if not ret:
        break
    gray = cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
    faces = face_classifier.detectMultiScale(gray,1.3,5)

    for (x,y,w,h) in faces:
        cv2.rectangle(frame,(x,y),(x+w,y+h),(255,0,0),2)
        roi_gray = gray[y:y+h,x:x+w]
        roi_gray = cv2.resize(roi_gray,(48,48),interpolation=cv2.INTER_AREA)
    # rect,face,image = face_detector(frame)


        if np.sum([roi_gray])!=0:
            roi = roi_gray.astype('float')/255.0
            roi = img_to_array(roi)
            roi = np.expand_dims(roi,axis=0)

        # make a prediction on the ROI, then lookup the class

            preds = classifier.predict(roi)[0]
            label=class_labels[preds.argmax()]
            label_position = (x,y)
            cv2.putText(frame,label,label_position,cv2.FONT_HERSHEY_SIMPLEX,2,(0,255,0),3)
        else:
            cv2.putText(frame,'No Face Found',(20,60),cv2.FONT_HERSHEY_SIMPLEX,2,(0,255,0),3)
    cv2.imshow('Emotion Detector',frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows() 




#--------------------------------------------------------------------------------------------------------------------------------------------------


import os
import tensorflow as tf
from keras import layers, models
from keras.preprocessing.image import ImageDataGenerator

# Define data directory and image dimensions
data_dir = r"C:\Users\bbhar\Downloads\Compressed\archive_5\Images"
img_width, img_height = 224, 224

# Create a data generator with data augmentation for training
train_datagen = ImageDataGenerator(
    rescale=1./255,
    rotation_range=30,
    shear_range=0.2,
    zoom_range=0.2,
    horizontal_flip=True,
    validation_split=0.1 # 20% of the data will be used for validation
)

# Define the data generator for training (including validation split)
train_generator = train_datagen.flow_from_directory(
    data_dir,
    target_size=(img_width, img_height),
    batch_size=40,
    class_mode='categorical',
    subset='training'  # Specify that this is the training set
)

# Define the data generator for validation
val_generator = train_datagen.flow_from_directory(
    data_dir,
    target_size=(img_width, img_height),
    batch_size=40,
    class_mode='categorical',
    subset='validation'  # Specify that this is the validation set
)

# Create a simple CNN model
model = models.Sequential([
    layers.Conv2D(32, (3, 3), activation='elu', input_shape=(img_width, img_height, 3)),
    layers.MaxPooling2D((2, 2)),
    layers.Conv2D(64, (3, 3), activation='elu'),
    layers.MaxPooling2D((2, 2)),
    layers.Dropout((0.2)),
    layers.Conv2D(128, (3, 3), activation='elu'),
    layers.MaxPooling2D((2, 2)),
    layers.Dropout((0.2)),
    layers.Conv2D(128, (3, 3), activation='elu'),
    layers.MaxPooling2D((2, 2)),
    layers.Dropout((0.2)),
    layers.Conv2D(128, (3, 3), activation='elu'),
    layers.MaxPooling2D((2, 2)),
    layers.Dropout((0.2)),
    layers.Flatten(),
    layers.Dense(128, activation='elu'),
    layers.Dropout(0.5),
    layers.Dense(len(train_generator.class_indices), activation='softmax')
])

# Compile the model
model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

# Train the model
history = model.fit(train_generator, epochs=25, validation_data=val_generator)

# Save the trained model
model.save('actor_recognition_model.h5')

# import numpy as np
# from keras.preprocessing import image
# from keras.models import load_model

# # Load the trained model
# model = load_model(r"F:\python\actor_recognition_model.h5")

# # Define the path to the input image you want to recognize
# input_image_path = r"C:\Users\bbhar\Downloads\Compressed\archive_5\HrithikRoshan_81.jpg"

# # Load and preprocess the input image
# img = image.load_img(input_image_path, target_size=(224, 224))
# img_array = image.img_to_array(img)
# img_array = np.expand_dims(img_array, axis=0)
# img_array /= 255.0  # Normalize the image

# # Make a prediction
# predictions = model.predict(img_array)

# # Get the predicted class label
# predicted_class = np.argmax(predictions)

# # Get the class indices and labels
# class_indices = train_generator.class_indices
# class_labels = list(class_indices.keys())

# # Get the actor name from the predicted class label
# actor_name = class_labels[predicted_class]

# # Print the result
# print(f"The input image is predicted to be the actor: {actor_name}")



#--------------------------------------------------------------------------------------------------------------------------------------------------



import requests
import os

def download_repos(token, org_or_user, target_folder):
    # Set the GitHub token
    headers = {"Authorization": f"token {token}"}

    # Get the list of repositories for the organization or user
    response = requests.get(f"https://api.github.com/users/{org_or_user}/repos", headers=headers)

    # Check if the request was successful
    if response.status_code == 200:
        repos = response.json()

        # Loop through the repositories and clone each one
        for repo in repos:
            clone_url = repo["clone_url"]
            repo_name = repo["name"]
            target_path = os.path.join(target_folder, repo_name)

            # Check if the repository already exists locally
            if os.path.exists(target_path):
                print(f"Repository '{repo_name}' already exists in '{target_folder}'. Skipping.")
            else:
                # Clone the repository into the specified target folder
                os.system(f"git clone {clone_url} {target_path}")
                print(f"Repository '{repo_name}' cloned successfully into '{target_folder}'.")
    else:
        print(f"Failed to retrieve repositories. Status code: {response.status_code}")

if __name__ == "__main__":
    # Replace 'YOUR_GITHUB_TOKEN' with your GitHub personal access token
    github_token = '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'

    # Replace 'ORG_OR_USER' with the GitHub organization or user name
    organization_or_user = 'nicknochnack'

    # Specify the target folder where repositories will be cloned
    target_folder = 'F:\\python\\Git_Hub\\nicknochnack'

    download_repos(github_token, organization_or_user, target_folder)




#--------------------------------------------------------------------------------------------------------------------------------------------------


import jsonlines
import tensorflow as tf
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
import numpy as np

# Function to create a data generator
def data_generator(sequences, targets, batch_size):
    steps = len(sequences) // batch_size
    while True:
        for i in range(steps):
            batch_sequences = sequences[i * batch_size: (i + 1) * batch_size]
            batch_targets = next(targets)
            yield np.array(batch_sequences), np.array(batch_targets)

file_path = "F:\\python\\data\\train.jsonl"

with jsonlines.open(file_path, mode='r') as reader:
    data = [line for line in reader]

# Extract text
text = [example["text"] for example in data]

# Tokenize text
tokenizer = Tokenizer()
tokenizer.fit_on_texts(text)

# Convert text to sequences
sequences = tokenizer.texts_to_sequences(text)

# Limit the sequence length
max_length = 50  # Adjust as needed
input_sequences = [seq[:max_length] for seq in sequences]
target_sequences = [seq[1:max_length] for seq in sequences]

# Pad sequences for fixed input length
input_padded = pad_sequences(input_sequences, maxlen=max_length, padding='post')
target_padded = pad_sequences(target_sequences, maxlen=max_length, padding='post')

# Function to create a generator for the target sequences
def target_generator(targets, batch_size):
    steps = len(targets) // batch_size
    while True:
        for i in range(steps):
            yield targets[i * batch_size: (i + 1) * batch_size]

# Convert target sequences to categorical using a generator
batch_size = 80
target_generator_obj = target_generator(target_padded, batch_size)
target_categorical_generator = (tf.keras.utils.to_categorical(batch, num_classes=len(tokenizer.word_index) + 1) for batch in target_generator_obj)

# Define the model with further reduced complexity
model = tf.keras.Sequential([
    tf.keras.layers.Embedding(input_dim=len(tokenizer.word_index) + 1, output_dim=50, input_length=max_length),
    tf.keras.layers.LSTM(50, return_sequences=True),
    tf.keras.layers.Dense(len(tokenizer.word_index) + 1, activation='softmax')
])

# Compile the model using mixed precision
optimizer = tf.keras.optimizers.Adam()
model.compile(optimizer=optimizer, loss='categorical_crossentropy', metrics=['accuracy'])

# Train the model using the data generator
steps_per_epoch = len(input_padded) // batch_size
model.fit(data_generator(input_padded, target_categorical_generator, batch_size),
          epochs=5, steps_per_epoch=steps_per_epoch)

model.save("path_to_save_model/my_text_generation_model.h5")

# Use the trained model for inference
test_prompt = "There are"
test_prompt_sequence = tokenizer.texts_to_sequences([test_prompt])
test_prompt_padded = pad_sequences(test_prompt_sequence, maxlen=max_length, padding='post')

# Initial prediction
predicted_sequence = model.predict(test_prompt_padded)

# Generate the next word iteratively
for _ in range(5):  # You can adjust the number of words to generate
    next_word_index = tf.argmax(predicted_sequence[0, -1, :]).numpy() - 1

    # Check if the index is valid
    if next_word_index < 0:
        break

    next_word = tokenizer.index_word[next_word_index]
    test_prompt += " " + next_word

    # Update the input sequence for the next prediction
    test_prompt_sequence = tokenizer.texts_to_sequences([test_prompt])
    test_prompt_padded = pad_sequences(test_prompt_sequence, maxlen=max_length, padding='post')

    # Predict the next word
    predicted_sequence = model.predict(test_prompt_padded)

print("Generated Text:", test_prompt)




#--------------------------------------------------------------------------------------------------------------------------------------------------



from transformers import pipeline

# Load pre-trained question-answering model
qa_pipeline = pipeline("question-answering")

# Example context
context = '''France, located in Western Europe, is known for its rich cultural heritage, diverse geography encompassing mountains, plains, and 
coastlines along the Atlantic Ocean and the Mediterranean Sea. The capital, Paris, is renowned for iconic landmarks like the Eiffel Tower and 
the Louvre Museum. The official language is French, reflecting the country's literary and cultural prominence. With a unitary semi-presidential 
republic system, France has played a pivotal role in shaping European history, from the Roman Empire to the Middle Ages and the French Revolution. 
Its contributions extend to art, literature, philosophy, and culinary arts, making French cuisine globally celebrated. Boasting a developed 
economy, membership in the European Union, and a prominent place in international politics, France remains a major tourist destination with a 
well-established education system and efficient transportation networks.'''

# Ask a question related to the context
question = "What did France played a pivotal role for?"

# Get the answer from the model
answer = qa_pipeline(question=question, context=context)

# Print the answer
print(answer)


#--------------------------------------------------------------------------------------------------------------------------------------------------


from __future__ import print_function
import keras
from keras.preprocessing.image import ImageDataGenerator
from keras.models import Sequential
from keras.layers import Dense,Flatten,Dropout,BatchNormalization,Activation
from keras.layers import Conv2D,MaxPooling2D
from keras.optimizers import RMSprop,SGD,Adam
from keras.callbacks import ModelCheckpoint, EarlyStopping, ReduceLROnPlateau


num_class = 5
img_row, img_col = (48,48)
batch_size = 45

train_data_dir = r"F:\python\pythonWork\data\archive\train"
validation_data_dir = r"F:\python\pythonWork\data\archive\test"

train_datagen = ImageDataGenerator(rescale = 1./255,
                                   rotation_range = 30,
                                   shear_range = 0.3,
                                   zoom_range = 0.3,
                                   fill_mode = 'nearest',
                                   horizontal_flip = True,
                                   width_shift_range = 0.3,
                                   height_shift_range = 0.3)

valid_datagen = ImageDataGenerator(rescale=1./255)

train_generator = train_datagen.flow_from_directory(train_data_dir,
                        target_size = (img_row,img_col),
                        class_mode = 'categorical',
                        batch_size= batch_size,
                        # classes=num_class,
                        shuffle=True,
                        color_mode='grayscale')

valid_generator = valid_datagen.flow_from_directory(validation_data_dir,
                        target_size = (img_row,img_col),
                        class_mode = 'categorical',
                        batch_size= batch_size,
                        # classes=num_class,
                        shuffle=True,
                        color_mode='grayscale')


model = Sequential()

# b1
model.add(Conv2D(32,(3,3),kernel_initializer = 'he_normal',padding = 'same', input_shape=(img_row,img_col,1)))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Conv2D(32,(3,3), padding = 'same', kernel_initializer = 'he_normal',input_shape = (img_row,img_col,1)))
model.add(Activation("elu"))
model.add(BatchNormalization())
model.add(MaxPooling2D(pool_size = (2,2)))
model.add(Dropout(0.2))

# b2
model.add(Conv2D(64,(3,3),kernel_initializer="he_normal",padding = 'same'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Conv2D(64,(3,3),kernel_initializer = 'he_normal',padding = 'same'))
model.add(Activation("elu"))
model.add(BatchNormalization())
model.add(MaxPooling2D(pool_size=(2,2)))
model.add(Dropout(0.2))

# b3
model.add(Conv2D(128,(3,3),kernel_initializer="he_normal",padding = 'same'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Conv2D(128,(3,3),kernel_initializer = 'he_normal',padding = 'same'))
model.add(Activation("elu"))
model.add(BatchNormalization())
model.add(MaxPooling2D(pool_size=(2,2)))
model.add(Dropout(0.2))

# b4
model.add(Conv2D(256,(3,3),kernel_initializer="he_normal",padding = 'same'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Conv2D(256,(3,3),kernel_initializer = 'he_normal',padding = 'same'))
model.add(Activation("elu"))
model.add(BatchNormalization())
model.add(MaxPooling2D(pool_size=(2,2)))
model.add(Dropout(0.2))

model.add(Flatten())
model.add(Dense(64, kernel_initializer = 'he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Dropout(0.5))


model.add(Dense(64, kernel_initializer = 'he_normal'))
model.add(Activation('elu'))
model.add(BatchNormalization())
model.add(Dropout(0.5))

model.add(Dense(num_class, kernel_initializer = 'he_normal'))
model.add(Activation('softmax'))

print(model.summary())


checkpoint = ModelCheckpoint('Emotion_vgg.h5',
                             monitor='val_loss',
                             mode='min',
                             save_best_only=True,
                             verbose=1)

earlystop = EarlyStopping(monitor='val_loss',
                          min_delta=0,
                          patience=3,
                          verbose=1,
                          restore_best_weights=True
                          )

reduce_lr = ReduceLROnPlateau(monitor='val_loss',
                              factor=0.2,
                              patience=3,
                              verbose=1,
                              min_delta=0.0001)

callbacks = [earlystop,checkpoint,reduce_lr]

model.compile(loss='categorical_crossentropy',
              optimizer = Adam(learning_rate=0.001),
              metrics=['accuracy'])

nb_train_samples = 48352
nb_validation_samples = 6043
epochs=5

history=model.fit(
                train_generator,
                steps_per_epoch=nb_train_samples//batch_size,
                epochs=epochs,
                callbacks=callbacks,
                validation_data=valid_generator,
                validation_steps=nb_validation_samples//batch_size)








#--------------------------------------------------------------------------------------------------------------------------------------------------



import streamlit as st
from langchain.prompts import PromptTemplate
from langchain.llms import CTransformers

## Function To get response from LLAma 2 model

def getLLamaresponse(input_text,no_words,blog_style):

    ### LLama2 model
    llm=CTransformers(model=r"C:\Users\bbhar\Downloads\llama-2-7b-chat.ggmlv3.q8_0.bin",
                      model_type='llama',
                      config={'max_new_tokens':256,
                              'temperature':0.01})
    
    ## Prompt Template

    template="""
        Write a blog for {blog_style} job profile for a topic {input_text}
        within {no_words} words.
            """
    
    prompt=PromptTemplate(input_variables=["blog_style","input_text",'no_words'],
                          template=template)
    
    ## Generate the ressponse from the LLama 2 model
    response=llm(prompt.format(blog_style=blog_style,input_text=input_text,no_words=no_words))
    print(response)
    return response






st.set_page_config(page_title="Generate Blogs",
                    page_icon='🤖',
                    layout='centered',
                    initial_sidebar_state='collapsed')

st.header("Generate Blogs 🤖")

input_text=st.text_input("Enter the Blog Topic")

## creating to more columns for additonal 2 fields

col1,col2=st.columns([5,5])

with col1:
    no_words=st.text_input('No of Words')
with col2:
    blog_style=st.selectbox('Writing the blog for',
                            ('Researchers','Data Scientist','Common People'),index=0)
    
submit=st.button("Generate")

## Final response
if submit:
    st.write(getLLamaresponse(input_text,no_words,blog_style))



#--------------------------------------------------------------------------------------------------------------------------------------------------


from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity

def calculate_cosine_similarity(sentence1, sentence2):
    # Create a CountVectorizer to convert text into token counts
    vectorizer = CountVectorizer()

    # Fit and transform the sentences into vectors
    vectors = vectorizer.fit_transform([sentence1, sentence2])

    # Calculate cosine similarity
    similarity_matrix = cosine_similarity(vectors)

    # The element [0, 1] or [1, 0] in the matrix represents the cosine similarity between the two sentences
    cosine_similarity_score = similarity_matrix[0, 1]

    return cosine_similarity_score

# Example usage
sentence1 = "Natural Language Processing is fascinating."
sentence2 = "NLP is fascinating"

similarity_score = calculate_cosine_similarity(sentence1, sentence2)

print(f"Cosine Similarity: {similarity_score}")




#--------------------------------------------------------------------------------------------------------------------------------------------------









#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------







#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------









#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------







#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------









#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------







#--------------------------------------------------------------------------------------------------------------------------------------------------






#--------------------------------------------------------------------------------------------------------------------------------------------------